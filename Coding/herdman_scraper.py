import json
import os
import re
import time
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEFAULT_MANAGER_ID = 53225
BASE_URL = "https://www.sofascore.com/api/v1"

# ── Selenium setup ────────────────────────────────────────────────────────────

def create_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()),
        options=options,
    )
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


def fetch_json(driver, url, retries=3):
    for attempt in range(retries):
        try:
            driver.get(url)
            time.sleep(0.8)
            raw = driver.find_element("tag name", "pre").text
            return json.loads(raw)
        except Exception as e:
            print(f"    Retry {attempt + 1} on {url} — {e}")
            time.sleep(2 ** attempt)
    return None

# ── parsers ───────────────────────────────────────────────────────────────────

def ts_to_date(ts):
    return datetime.utcfromtimestamp(ts).strftime("%Y-%m-%d") if ts else ""

def winner_label(code):
    return {1: "Home", 2: "Away", 3: "Draw"}.get(code, "")

def parse_lineups(data):
    if not data:
        return {}, {}, {}
    formations, starters, subs = {}, {}, {}
    for side in ["home", "away"]:
        sd = data.get(side, {})
        formations[f"{side}_formation"] = sd.get("formation", "")
        starters[side] = [
            f"{p.get('jerseyNumber', '?')}. {p.get('player', {}).get('name', '?')} ({p.get('position', '')})"
            for p in sd.get("players", [])
            if not p.get("substitute", True)
        ]
        subs[side] = [
            f"{p.get('jerseyNumber', '?')}. {p.get('player', {}).get('name', '?')}"
            for p in sd.get("players", [])
            if p.get("substitute", True)
        ]
    return formations, starters, subs

def parse_statistics(data):
    flat = {}
    if not data:
        return flat
    for period_block in data.get("statistics", []):
        period = period_block.get("period", "ALL")
        for group in period_block.get("groups", []):
            for item in group.get("statisticsItems", []):
                name = item["name"]
                flat[(period, name, "home")] = item.get("home", "")
                flat[(period, name, "away")] = item.get("away", "")
    return flat

# ── Excel helpers ─────────────────────────────────────────────────────────────

def header_style(cell, bg="1F3864", fg="FFFFFF"):
    cell.font = Font(name="Arial", bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border

def result_row_color(winner_code):
    return {1: "E2EFDA", 2: "FCE4D6", 3: "FFF2CC"}.get(winner_code, "FFFFFF")

def plain_cell(cell, value, center=False, row_bg=None):
    cell.value = value
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    if row_bg:
        cell.fill = PatternFill("solid", start_color=row_bg)

# ── main ──────────────────────────────────────────────────────────────────────

def lookup_manager_name(manager_id: int) -> str:
    """Spin up a brief driver session just to resolve the manager's display name."""
    driver = create_driver()
    try:
        data = fetch_json(driver, f"{BASE_URL}/manager/{manager_id}")
        return (data or {}).get("manager", {}).get("name", f"manager_{manager_id}")
    finally:
        driver.quit()


def run_scraper(manager_id: int, log_fn=print) -> str:
    """
    Run the full match scraper for *manager_id*.
    Calls log_fn(msg) for progress updates.
    Returns the absolute path to the saved Excel file.
    """
    log_fn("=" * 55)
    log_fn("  Manager Match Scraper  (Selenium)")
    log_fn("=" * 55)

    driver = create_driver()

    # 0. fetch manager info
    log_fn("\n[0/3] Fetching manager info...")
    manager_data = fetch_json(driver, f"{BASE_URL}/manager/{manager_id}")
    manager_name = (manager_data or {}).get("manager", {}).get("name", f"manager_{manager_id}")
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", manager_name)
    log_fn(f"      Manager: {manager_name}\n")

    # 1. collect all events
    log_fn("[1/3] Fetching match list...")
    all_events = []
    page = 0
    while True:
        data = fetch_json(driver, f"{BASE_URL}/manager/{manager_id}/events/last/{page}")
        events = (data or {}).get("events", [])
        if not events:
            break
        all_events.extend(events)
        log_fn(f"      Page {page}: {len(events)} matches")
        page += 1
        time.sleep(1)
    log_fn(f"      Total: {len(all_events)} matches\n")

    # 2. enrich each event with lineups + stats
    log_fn("[2/3] Fetching lineups & statistics...")
    records = []
    all_stat_keys = set()

    for i, event in enumerate(all_events):
        eid = event["id"]
        home = event.get("homeTeam", {}).get("name", "?")
        away = event.get("awayTeam", {}).get("name", "?")
        log_fn(f"  [{i + 1:>3}/{len(all_events)}] {home} vs {away}  (id:{eid})")

        lineup_data = fetch_json(driver, f"{BASE_URL}/event/{eid}/lineups")
        time.sleep(1)
        stats_data  = fetch_json(driver, f"{BASE_URL}/event/{eid}/statistics")
        time.sleep(1.2)

        formations, starters, subs = parse_lineups(lineup_data)
        stats = parse_statistics(stats_data)
        all_stat_keys.update(stats.keys())

        records.append({
            "event_id":       eid,
            "date":           ts_to_date(event.get("startTimestamp")),
            "competition":    event.get("tournament", {}).get("name", ""),
            "season":         event.get("season", {}).get("name", ""),
            "home_team":      home,
            "away_team":      away,
            "home_score":     event.get("homeScore", {}).get("current", ""),
            "away_score":     event.get("awayScore", {}).get("current", ""),
            "ht_home":        event.get("homeScore", {}).get("period1", ""),
            "ht_away":        event.get("awayScore", {}).get("period1", ""),
            "winner_code":    event.get("winnerCode"),
            "result":         winner_label(event.get("winnerCode")),
            "home_formation": formations.get("home_formation", ""),
            "away_formation": formations.get("away_formation", ""),
            "home_starters":  " | ".join(starters.get("home", [])),
            "away_starters":  " | ".join(starters.get("away", [])),
            "home_subs":      " | ".join(subs.get("home", [])),
            "away_subs":      " | ".join(subs.get("away", [])),
            "stats":          stats,
        })

    driver.quit()

    # sort stat columns: ALL → 1ST → 2ND, then alphabetically
    periods_order = ["ALL", "1ST", "2ND"]
    stat_names_by_period = {}
    for (period, stat_name, side) in all_stat_keys:
        stat_names_by_period.setdefault(period, set()).add(stat_name)

    sorted_stat_keys = []
    for period in periods_order:
        for stat_name in sorted(stat_names_by_period.get(period, [])):
            for side in ["home", "away"]:
                sorted_stat_keys.append((period, stat_name, side))

    # 3. build Excel workbook
    log_fn("\n[3/3] Building Excel workbook...")
    wb = Workbook()

    # ── Sheet 1: Match Results ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Match Results"
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 30

    results_cols = [
        ("Match ID",       10),
        ("Date",           12),
        ("Competition",    22),
        ("Season",         18),
        ("Home Team",      20),
        ("Score",           9),
        ("Away Team",      20),
        ("HT Score",        9),
        ("Result",          9),
        ("Home Formation", 15),
        ("Away Formation", 15),
    ]
    for ci, (name, width) in enumerate(results_cols, 1):
        header_style(ws1.cell(row=1, column=ci, value=name))
        ws1.column_dimensions[get_column_letter(ci)].width = width

    for ri, rec in enumerate(records, 2):
        bg = result_row_color(rec["winner_code"])
        row_vals = [
            rec["event_id"], rec["date"], rec["competition"], rec["season"],
            rec["home_team"],
            f"{rec['home_score']} - {rec['away_score']}",
            rec["away_team"],
            f"{rec['ht_home']} - {rec['ht_away']}",
            rec["result"],
            rec["home_formation"], rec["away_formation"],
        ]
        for ci, val in enumerate(row_vals, 1):
            plain_cell(ws1.cell(row=ri, column=ci), val, center=True, row_bg=bg)

    apply_border(ws1, 1, len(records) + 1, 1, len(results_cols))
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(results_cols))}1"

    # ── Sheet 2: Lineups ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Lineups")
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 30

    lineup_cols = [
        ("Match ID",        10), ("Date",            12),
        ("Home Team",       20), ("Away Team",        20),
        ("Home Formation",  15), ("Away Formation",   15),
        ("Home Starting XI",60), ("Away Starting XI", 60),
        ("Home Subs",       45), ("Away Subs",        45),
    ]
    for ci, (name, width) in enumerate(lineup_cols, 1):
        header_style(ws2.cell(row=1, column=ci, value=name))
        ws2.column_dimensions[get_column_letter(ci)].width = width

    for ri, rec in enumerate(records, 2):
        row_vals = [
            rec["event_id"], rec["date"],
            rec["home_team"], rec["away_team"],
            rec["home_formation"], rec["away_formation"],
            rec["home_starters"], rec["away_starters"],
            rec["home_subs"],     rec["away_subs"],
        ]
        for ci, val in enumerate(row_vals, 1):
            plain_cell(ws2.cell(row=ri, column=ci), val)
        ws2.row_dimensions[ri].height = 45

    apply_border(ws2, 1, len(records) + 1, 1, len(lineup_cols))

    # ── Sheet 3: Statistics ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Statistics")
    ws3.freeze_panes = "E2"
    ws3.row_dimensions[1].height = 55

    fixed_cols = [
        ("Match ID", 10), ("Date", 12), ("Home Team", 20), ("Away Team", 20),
    ]
    for ci, (name, width) in enumerate(fixed_cols, 1):
        header_style(ws3.cell(row=1, column=ci, value=name))
        ws3.column_dimensions[get_column_letter(ci)].width = width

    period_colors = {"ALL": "1F3864", "1ST": "375623", "2ND": "7F3F98"}
    col_map = {}
    current_col = len(fixed_cols) + 1

    for (period, stat_name, side) in sorted_stat_keys:
        label = f"[{period}]\n{stat_name}\n({side.title()})"
        cell = ws3.cell(row=1, column=current_col, value=label)
        header_style(cell, bg=period_colors.get(period, "1F3864"))
        ws3.column_dimensions[get_column_letter(current_col)].width = 18
        col_map[(period, stat_name, side)] = current_col
        current_col += 1

    for ri, rec in enumerate(records, 2):
        for ci, val in enumerate(
            [rec["event_id"], rec["date"], rec["home_team"], rec["away_team"]], 1
        ):
            plain_cell(ws3.cell(row=ri, column=ci), val, center=True)

        zebra = "F2F2F2" if ri % 2 == 0 else None
        for (period, stat_name, side), col_idx in col_map.items():
            val = rec["stats"].get((period, stat_name, side), "")
            cell = ws3.cell(row=ri, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if zebra:
                cell.fill = PatternFill("solid", start_color=zebra)

    apply_border(ws3, 1, len(records) + 1, 1, current_col - 1)
    ws3.auto_filter.ref = f"A1:{get_column_letter(current_col - 1)}1"

    # save
    output_dir = os.path.join("manager_output", safe_name)
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"{safe_name}_matches.xlsx")
    wb.save(output_path)
    log_fn(f"\n  Saved → {output_path}")
    log_fn(f"  Matches  : {len(records)}")
    log_fn(f"  Stat cols: {current_col - len(fixed_cols) - 1}")
    log_fn("\nDone!")
    return os.path.abspath(output_path)


def main():
    run_scraper(DEFAULT_MANAGER_ID)


if __name__ == "__main__":
    main()
