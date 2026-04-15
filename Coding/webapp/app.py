import sys
import os
import threading
import uuid
sys.path.insert(0, os.path.dirname(__file__))
# Coding/ dir — needed for scraper_cli, driver, helpers, heatmap
_CODING_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), '..'))
if _CODING_DIR not in sys.path:
    sys.path.insert(0, _CODING_DIR)
# Workspace root — needed for tactical_match_engine package
_WORKSPACE_ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', '..'))
if _WORKSPACE_ROOT not in sys.path:
    sys.path.insert(0, _WORKSPACE_ROOT)

from flask import Flask, render_template, request, jsonify, send_file
from data_loader import (
    get_all_competitions, get_competition, get_club, get_club_players,
    get_player, search_players, get_all_players_flat, get_all_clubs_flat,
    get_heatmap_path, get_home_stats
)
import csv
import io
import os

# ── Path resolution (works both in dev and in the frozen .exe) ────────────────
def _resolve_dir(env_key: str, *rel_parts) -> str:
    """Return the env-var value if set, otherwise resolve relative to __file__."""
    from_env = os.environ.get(env_key)
    if from_env:
        return from_env
    return os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), *rel_parts))


_WEBAPP_DIR = _resolve_dir('TACTICAL_WEBAPP_DIR')
OUTPUT_DIR  = _resolve_dir('TACTICAL_OUTPUT_DIR', '..', 'output')

app = Flask(
    __name__,
    template_folder=os.path.join(_WEBAPP_DIR, 'templates'),
    static_folder=os.path.join(_WEBAPP_DIR, 'static'),
)

# ── Home ──────────────────────────────────────────────────────────────────
@app.route('/')
def home():
    stats = get_home_stats(OUTPUT_DIR)
    return render_template('home.html', stats=stats)

# ── Competitions: League Overview ──────────────────────────────────
@app.route('/competitions')
def index():
    competitions = get_all_competitions(OUTPUT_DIR)
    return render_template('index.html', competitions=competitions)

# ── Competition detail ────────────────────────────────────────────────────────
@app.route('/competition/<country>/<competition>')
def competition(country, competition):
    data = get_competition(OUTPUT_DIR, country, competition)
    from tactical_match_engine.services.json_loader import get_league_intensity
    cvs_score = round(get_league_intensity(country.replace('_', ' '), competition.replace('_', ' ')) * 100, 1)
    return render_template('competition.html', data=data, country=country, competition=competition, cvs_score=cvs_score)

# ── Club overview ─────────────────────────────────────────────────────────────
@app.route('/competition/<country>/<competition>/<club>')
def club(country, competition, club):
    data    = get_club(OUTPUT_DIR, country, competition, club)
    players = get_club_players(OUTPUT_DIR, country, competition, club)
    return render_template('club.html', data=data, players=players,
                           country=country, competition=competition, club=club)

# ── Player profile ────────────────────────────────────────────────────────────
@app.route('/player/<country>/<competition>/<club>/<player_file>')
def player(country, competition, club, player_file):
    data         = get_player(OUTPUT_DIR, country, competition, club, player_file)
    heatmap_path = get_heatmap_path(OUTPUT_DIR, country, competition, club, player_file)
    from tactical_match_engine.services.json_loader import get_league_intensity
    cvs_score = round(get_league_intensity(country.replace('_', ' '), competition.replace('_', ' ')) * 100, 1)
    # Build player path for quick-links
    player_path = f"{country}/{competition}/{club}/{player_file}"
    # Match data enrichment
    player_id = data.get('player_id', '')
    match_per90  = None
    match_count  = 0
    if player_id:
        try:
            _appearances = get_player_match_history(int(player_id), MATCH_OUTPUT_DIR)
            match_count  = len(_appearances)
            if match_count > 0:
                match_per90 = compute_per90(_appearances)
        except Exception:
            pass
    return render_template('player.html', data=data, heatmap=heatmap_path,
                           country=country, competition=competition, club=club,
                           cvs_score=cvs_score, player_path=player_path,
                           match_per90=match_per90, match_count=match_count)

# ── Serve heatmap images ──────────────────────────────────────────────────────
@app.route('/heatmap/<path:filepath>')
def heatmap_image(filepath):
    full = os.path.join(OUTPUT_DIR, filepath)
    if os.path.exists(full):
        return send_file(full, mimetype='image/png')
    return '', 404

# ── Player search & filter ────────────────────────────────────────────────────
@app.route('/players')
def players():
    competitions = get_all_competitions(OUTPUT_DIR)
    return render_template('players.html', competitions=competitions)

@app.route('/api/players')
def api_players():
    filters = {
        'competition': request.args.get('competition', ''),
        'club':        request.args.get('club', ''),
        'position':    request.args.get('position', ''),
        'specific_pos': request.args.get('specific_pos', ''),
        'nationality': request.args.get('nationality', ''),
        'age_min':     request.args.get('age_min', ''),
        'age_max':     request.args.get('age_max', ''),
        'apps_min':    request.args.get('apps_min', ''),
        'rating_min':  request.args.get('rating_min', ''),
        'goals_min':   request.args.get('goals_min', ''),
        'assists_min': request.args.get('assists_min', ''),
        # Advanced
        'pass_acc_min':    request.args.get('pass_acc_min', ''),
        'duels_min':       request.args.get('duels_min', ''),
        'aerial_min':      request.args.get('aerial_min', ''),
        'dribbles_min':    request.args.get('dribbles_min', ''),
        'tackles_min':     request.args.get('tackles_min', ''),
        'key_passes_min':  request.args.get('key_passes_min', ''),
        'sort_by':         request.args.get('sort_by', 'rating'),
        'sort_dir':        request.args.get('sort_dir', 'desc'),
        'page':            int(request.args.get('page', 1)),
        'per_page':        int(request.args.get('per_page', 50)),
    }
    result = search_players(OUTPUT_DIR, filters)
    return jsonify(result)
# ── Tactical Compatibility ────────────────────────────────────────────────────
@app.route('/compatibility')
def compatibility():
    return render_template('compatibility.html')

@app.route('/api/compatibility')
def api_compatibility():
    player_path = request.args.get('player', '')
    role_file   = request.args.get('role', '')

    if not player_path or not role_file:
        return jsonify({'error': 'player and role are required'}), 400

    parts = player_path.split('/')
    if len(parts) != 4:
        return jsonify({'error': 'invalid player path'}), 400

    full_path = os.path.join(
        OUTPUT_DIR, parts[0], parts[1], parts[2], 'Players',
        parts[3] if parts[3].endswith('.json') else parts[3] + '.json'
    )
    if not os.path.exists(full_path):
        return jsonify({'error': 'player file not found'}), 404

    try:
        from tactical_match_engine.services.json_loader import load_sofascore_player
        from tactical_match_engine.engine.role_encoder import get_role_fitness_vector

        player_data = load_sofascore_player(full_path)
        result = get_role_fitness_vector(
            player_data['raw_stats'],
            player_data['position'],
            role_path=role_file
        )
        if result is None:
            return jsonify({'error': 'no role profile for this position'}), 404

        return jsonify({
            'player_name':       player_data['name'],
            'player_position':   player_data['position'],
            'player_positions':  player_data.get('positions', []),
            'player_age':        player_data['age'],
            'player_team':       player_data.get('stats', {}).get('team', parts[2].replace('_', ' ')),
            'league_intensity':  round(player_data['current_league_intensity'], 4),
            'role_name':         result['role_profile']['position'],
            'overall_score':     result['overall_score'],
            'category_scores':   result['category_scores'],
            'weights':           result['weights'],
            'metric_details':    result['metric_details'],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Compatibility league average ──────────────────────────────────────────────

# Broad position codes that belong to each position line
_LINE_BROAD_POS = {
    '1. Back-Line':  {'DC', 'DL', 'DR', 'WBL', 'WBR', 'CB'},
    '2. Mid-Line':   {'DM', 'MC', 'AM', 'ML', 'MR', 'CM'},
    '3. Front-Line': {'LW', 'RW', 'ST', 'CF', 'SS', 'FW'},
}

@app.route('/api/compatibility_avg')
def api_compatibility_avg():
    """
    Score every position-matched player in the selected player's league against
    the chosen role profile and return the averaged result.

    Query params:
      player – "Country/Competition/Club/filename.json"  (derives league + self-exclude)
      role   – role profile path (e.g. "2. Mid-Line/I_Central_Midfielder.json")
    """
    player_path = request.args.get('player', '').strip()
    role_file   = request.args.get('role', '').strip()

    if not player_path or not role_file:
        return jsonify({'error': 'player and role are required'}), 400

    parts = player_path.split('/')
    if len(parts) != 4:
        return jsonify({'error': 'invalid player path'}), 400

    country, competition, club_slug, filename = parts
    target_file = filename if filename.endswith('.json') else filename + '.json'

    # Infer which position line this role belongs to from the path prefix
    role_line  = role_file.split('/')[0]           # e.g. "1. Back-Line"
    line_codes = _LINE_BROAD_POS.get(role_line, set())

    try:
        import glob as _glob
        from tactical_match_engine.services.json_loader import load_sofascore_player
        from tactical_match_engine.engine.role_encoder import get_role_fitness_vector

        comp_dir = os.path.join(OUTPUT_DIR, country, competition)
        if not os.path.isdir(comp_dir):
            return jsonify({'error': 'competition directory not found'}), 404

        pool_results = []

        for club_name in sorted(os.listdir(comp_dir)):
            full_club = os.path.join(comp_dir, club_name)
            if not os.path.isdir(full_club):
                continue
            players_dir = os.path.join(full_club, 'Players')
            if not os.path.isdir(players_dir):
                continue
            for fpath in sorted(_glob.glob(os.path.join(players_dir, '*.json'))):
                # Exclude the selected player from their own pool
                if club_name == club_slug and os.path.basename(fpath) == target_file:
                    continue
                try:
                    pd  = load_sofascore_player(fpath)
                    pos = [str(p).upper() for p in pd.get('positions', [pd.get('position', '')])]
                    if line_codes and not any(p in line_codes for p in pos):
                        continue
                    res = get_role_fitness_vector(pd['raw_stats'], pd['position'], role_path=role_file)
                    if res is None:
                        continue
                    pool_results.append(res)
                except Exception:
                    continue

        pool_size = len(pool_results)

        if pool_size == 0:
            return jsonify({
                'avg_pool_size':   0,
                'overall_score':   0,
                'category_scores': {},
                'weights':         {},
                'metric_details':  [],
            })

        avg_overall = round(sum(r['overall_score'] for r in pool_results) / pool_size, 2)

        cat_buckets = {}
        for r in pool_results:
            for cat, sc in r['category_scores'].items():
                cat_buckets.setdefault(cat, []).append(sc)
        avg_cats = {cat: round(sum(v) / len(v), 2) for cat, v in cat_buckets.items()}

        metric_score_buckets = {}
        metric_raw_buckets   = {}
        for r in pool_results:
            for m in r['metric_details']:
                metric_score_buckets.setdefault(m['name'], []).append(m['score'])
                metric_raw_buckets.setdefault(m['name'],   []).append(m['raw_value'])

        avg_metrics = [
            {
                'name':             m['name'],
                'category':         m['category'],
                'higher_is_better': m['higher_is_better'],
                'raw_value':        round(sum(metric_raw_buckets[m['name']])   / len(metric_raw_buckets[m['name']]),   4),
                'score':            round(sum(metric_score_buckets[m['name']]) / len(metric_score_buckets[m['name']]), 2),
            }
            for m in pool_results[0]['metric_details']
        ]

        return jsonify({
            'avg_pool_size':   pool_size,
            'overall_score':   avg_overall,
            'category_scores': avg_cats,
            'weights':         pool_results[0]['weights'],
            'metric_details':  avg_metrics,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Club Compatibility ────────────────────────────────────────────────────────

# Reverse-lookup: role profile path → set of position codes that play that role.
# Built from role_encoder.POSITION_CODE_TO_ROLE_FILE at import time.
def _build_role_pos_codes():
    from tactical_match_engine.engine.role_encoder import POSITION_CODE_TO_ROLE_FILE
    mapping = {}
    for code, rel_path in POSITION_CODE_TO_ROLE_FILE.items():
        mapping.setdefault(rel_path, set()).add(code.upper())
    return mapping

_ROLE_TO_POS_CODES = None  # lazy-initialised


def _pos_codes_for_role(role_path):
    """Return the set of position codes (e.g. {'DC'}) whose default role is role_path."""
    global _ROLE_TO_POS_CODES
    if _ROLE_TO_POS_CODES is None:
        _ROLE_TO_POS_CODES = _build_role_pos_codes()
    # Normalise slashes for lookup
    normalised = role_path.replace('\\', '/')
    return _ROLE_TO_POS_CODES.get(normalised, set())


# Position-line lookup used for position-filtered squad pool construction
POSITION_LINE = {
    "GK": "Back-Line",  "DC": "Back-Line",  "DL": "Back-Line",
    "DR": "Back-Line",  "WBL": "Back-Line", "WBR": "Back-Line",
    "DM": "Mid-Line",   "MC": "Mid-Line",   "ML": "Mid-Line",
    "MR": "Mid-Line",   "AM": "Mid-Line",
    "LW": "Front-Line", "RW": "Front-Line", "ST": "Front-Line",
    "SS": "Front-Line",
}

def get_position_line(positions: list) -> str:
    """Return the dominant position line (Back-Line / Mid-Line / Front-Line) for a position list."""
    counts = {}
    for pos in positions:
        line = POSITION_LINE.get(str(pos).upper())
        if line:
            counts[line] = counts.get(line, 0) + 1
    if not counts:
        return "Unknown"
    keys = list(counts.keys())
    return max(counts, key=lambda k: (counts[k], -keys.index(k)))


def _get_squad_role_analysis(country, competition, club, role_path, cand_line=None, exclude_path=None):
    """
    Load every player at *club* whose position line matches *cand_line*,
    score each against *role_path*, and return aggregated squad stats.
    Pass *exclude_path* (the candidate's own file path) to exclude them from the pool.
    Falls back to role-code matching when *cand_line* is None.
    """
    import glob as _glob
    from tactical_match_engine.services.json_loader import load_sofascore_player
    from tactical_match_engine.engine.role_encoder import get_role_fitness_vector

    players_dir = os.path.join(OUTPUT_DIR, country, competition, club, 'Players')
    empty = {'players': [], 'avg_overall': 0.0,
             'avg_category_scores': {}, 'avg_metric_scores': {}, 'avg_metric_raw': {}, 'player_count': 0}
    if not os.path.isdir(players_dir):
        return empty

    role_codes = None if cand_line else _pos_codes_for_role(role_path)

    squad = []
    for fpath in sorted(_glob.glob(os.path.join(players_dir, '*.json'))):
        if exclude_path and os.path.normpath(fpath) == os.path.normpath(exclude_path):
            continue
        try:
            pd = load_sofascore_player(fpath)
            player_positions = [str(p).upper() for p in pd.get('positions', [pd.get('position', '')])]
            if cand_line:
                if get_position_line(player_positions) != cand_line:
                    continue
            else:
                if role_codes and not any(p in role_codes for p in player_positions):
                    continue
            res = get_role_fitness_vector(pd['raw_stats'], pd['position'], role_path=role_path, flat_weights=True)
            if res is None:
                continue
            squad.append({
                'name':            pd['name'],
                'position':        pd['position'],
                'positions':       pd.get('positions', []),
                'age':             pd['age'],
                'overall_score':   res['overall_score'],
                'category_scores': res['category_scores'],
                'metric_details':  res['metric_details'],
                'rating':          round(float(pd['raw_stats'].get('rating', 0) or 0), 2),
                'appearances':     int(pd['raw_stats'].get('appearances', 0) or 0),
            })
        except Exception:
            continue

    if not squad:
        return empty

    avg_overall = sum(p['overall_score'] for p in squad) / len(squad)

    cat_buckets = {}
    for p in squad:
        for cat, score in p['category_scores'].items():
            cat_buckets.setdefault(cat, []).append(score)
    avg_cat = {cat: round(sum(v) / len(v), 2) for cat, v in cat_buckets.items()}

    metric_score_buckets = {}
    metric_raw_buckets   = {}
    for p in squad:
        for m in p['metric_details']:
            metric_score_buckets.setdefault(m['name'], []).append(m['score'])
            metric_raw_buckets.setdefault(m['name'],  []).append(m['raw_value'])
    avg_metric_scores = {k: round(sum(v) / len(v), 2) for k, v in metric_score_buckets.items()}
    avg_metric_raw    = {k: round(sum(v) / len(v), 4) for k, v in metric_raw_buckets.items()}

    return {
        'players':             sorted(squad, key=lambda x: x['overall_score'], reverse=True),
        'avg_overall':         round(avg_overall, 2),
        'avg_category_scores': avg_cat,
        'avg_metric_scores':   avg_metric_scores,
        'avg_metric_raw':      avg_metric_raw,
        'player_count':        len(squad),
    }


@app.route('/club_compatibility')
def club_compatibility():
    return render_template('club_compatibility.html')


@app.route('/api/club_search')
def api_club_search():
    q      = request.args.get('q', '').lower()
    league = request.args.get('league', '')   # optional "Country/Competition" filter
    result = get_all_clubs_flat(OUTPUT_DIR, query=q, limit=20, league_filter=league)
    return jsonify(result)


# ── League Suitability ────────────────────────────────────────────────────────
@app.route('/api/league_suitability')
def api_league_suitability():
    """
    Stage 1 of the two-step Wizard.
    Score every scraped league for a player+role combination.
    suitability_score = 0.6 * role_fitness + 0.4 * league_adaptation
    league_adaptation uses the sigmoid-based formula from physical_adaptation.py
    so upward leaps are penalised correctly (asymmetric, nonlinear).
    """
    player_path = request.args.get('player', '').strip()
    role_file   = request.args.get('role',   '').strip()

    if not player_path or not role_file:
        return jsonify({'error': 'player and role are required'}), 400

    p_parts = player_path.split('/')
    if len(p_parts) != 4:
        return jsonify({'error': 'invalid player path'}), 400

    player_fname = p_parts[3] if p_parts[3].endswith('.json') else p_parts[3] + '.json'
    player_full  = os.path.join(OUTPUT_DIR, p_parts[0], p_parts[1], p_parts[2], 'Players', player_fname)
    if not os.path.exists(player_full):
        return jsonify({'error': 'player file not found'}), 404

    try:
        import glob as _glob
        import json as _json
        from tactical_match_engine.services.json_loader import (
            load_sofascore_player, get_league_info
        )
        from tactical_match_engine.engine.role_encoder import get_role_fitness_vector
        from tactical_match_engine.engine.physical_adaptation import calculate_physical_adaptation

        cand        = load_sofascore_player(player_full)
        cand_result = get_role_fitness_vector(
            cand['raw_stats'], cand['position'], role_path=role_file, flat_weights=True
        )
        if cand_result is None:
            return jsonify({'error': 'no role profile available for this position'}), 404

        role_fitness          = cand_result['overall_score']
        player_intensity      = cand['current_league_intensity']
        player_global_rank    = cand.get('current_league_global_rank')
        player_league_name    = cand.get('current_league_name', '')
        role_name             = cand_result.get('role_profile', {}).get('position', role_file)
        cand_positions        = [str(p).upper() for p in cand.get('positions', [cand['position']])]
        cand_line             = get_position_line(cand_positions)
        age                   = cand.get('age', 26)

        leagues = []
        for country_entry in sorted(os.scandir(OUTPUT_DIR), key=lambda e: e.name):
            if not country_entry.is_dir():
                continue
            for comp_entry in sorted(os.scandir(country_entry.path), key=lambda e: e.name):
                if not comp_entry.is_dir():
                    continue

                t_info           = get_league_info(
                    country_entry.name.replace('_', ' '),
                    comp_entry.name.replace('_', ' ')
                )
                target_intensity = t_info['intensity']
                target_rank      = t_info.get('global_rank')

                player_count = 0
                for _pf in _glob.glob(os.path.join(comp_entry.path, '*', 'Players', '*.json')):
                    try:
                        with open(_pf, encoding='utf-8') as _f:
                            _pd = _json.load(_f)
                        _pos = [str(p).upper() for p in (_pd.get('positions') or [_pd.get('position', '')])]
                        if get_position_line(_pos) == cand_line:
                            player_count += 1
                    except Exception:
                        pass

                league_adaptation = round(
                    calculate_physical_adaptation(player_intensity, target_intensity) * 100.0, 2
                )
                suitability = round(0.6 * role_fitness + 0.4 * league_adaptation, 2)

                # Tier gap label
                if player_global_rank and target_rank:
                    rank_diff = player_global_rank - target_rank   # positive = target is stronger
                    if rank_diff > 5:
                        move_label = f'+{rank_diff} ranks up'
                    elif rank_diff < -5:
                        move_label = f'{abs(rank_diff)} ranks down'
                    else:
                        move_label = 'lateral move'
                else:
                    move_label = ''

                leagues.append({
                    'league':            comp_entry.name.replace('_', ' '),
                    'league_key':        f"{country_entry.name}/{comp_entry.name}",
                    'country':           country_entry.name.replace('_', ' '),
                    'suitability_score': suitability,
                    'league_intensity':  round(target_intensity, 4),
                    'global_rank':       target_rank,
                    'role_fitness':      round(role_fitness, 2),
                    'league_adaptation': league_adaptation,
                    'move_label':        move_label,
                    'player_count':      player_count,
                })

        leagues.sort(key=lambda x: x['suitability_score'], reverse=True)
        return jsonify({
            'player_name':            cand['name'],
            'role_name':              role_name,
            'player_league_name':     player_league_name,
            'player_global_rank':     player_global_rank,
            'player_league_intensity': round(player_intensity, 4),
            'leagues':                leagues,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Club Suitability ──────────────────────────────────────────────────────────
@app.route('/api/club_suitability')
def api_club_suitability():
    """
    Stage 2 of the two-step Wizard.
    Score every club in a selected league for a player+role combination.
    Reuses the existing club-fit scoring formula from /api/club_compatibility.
    """
    player_path = request.args.get('player', '').strip()
    league      = request.args.get('league', '').strip()   # "Country/Competition"
    role_file   = request.args.get('role',   '').strip()

    if not player_path or not league or not role_file:
        return jsonify({'error': 'player, league, and role are required'}), 400

    p_parts = player_path.split('/')
    l_parts = league.split('/')
    if len(p_parts) != 4:
        return jsonify({'error': 'invalid player path'}), 400
    if len(l_parts) != 2:
        return jsonify({'error': 'league must be Country/Competition'}), 400

    player_fname = p_parts[3] if p_parts[3].endswith('.json') else p_parts[3] + '.json'
    player_full  = os.path.join(OUTPUT_DIR, p_parts[0], p_parts[1], p_parts[2], 'Players', player_fname)
    if not os.path.exists(player_full):
        return jsonify({'error': 'player file not found'}), 404

    tc, tcomp    = l_parts
    league_dir   = os.path.join(OUTPUT_DIR, tc, tcomp)
    if not os.path.isdir(league_dir):
        return jsonify({'error': 'league directory not found'}), 404

    try:
        import json as _json
        from tactical_match_engine.services.json_loader import load_sofascore_player, get_league_info
        from tactical_match_engine.engine.role_encoder import get_role_fitness_vector
        from tactical_match_engine.engine.physical_adaptation import calculate_physical_adaptation

        cand        = load_sofascore_player(player_full)
        cand_result = get_role_fitness_vector(
            cand['raw_stats'], cand['position'], role_path=role_file, flat_weights=True
        )
        if cand_result is None:
            return jsonify({'error': 'no role profile available for this position'}), 404

        role_fitness       = cand_result['overall_score']
        player_intensity   = cand['current_league_intensity']
        player_global_rank = cand.get('current_league_global_rank')
        age                = cand.get('age', 26)

        t_info           = get_league_info(tc.replace('_', ' '), tcomp.replace('_', ' '))
        target_intensity = t_info['intensity']
        target_rank      = t_info.get('global_rank')
        league_adaptation = round(
            calculate_physical_adaptation(player_intensity, target_intensity) * 100.0, 2
        )

        # Tier gap label
        if player_global_rank and target_rank:
            rank_diff = player_global_rank - target_rank
            if rank_diff > 5:
                move_label = f'+{rank_diff} ranks up'
            elif rank_diff < -5:
                move_label = f'{abs(rank_diff)} ranks down'
            else:
                move_label = 'lateral move'
        else:
            move_label = ''

        cand_positions = [str(p).upper() for p in cand.get('positions', [cand['position']])]
        cand_line      = get_position_line(cand_positions)

        clubs = []
        for club_entry in sorted(os.scandir(league_dir), key=lambda e: e.name):
            if not club_entry.is_dir():
                continue

            squad     = _get_squad_role_analysis(tc, tcomp, club_entry.name, role_file, cand_line=cand_line)
            squad_avg = squad['avg_overall']
            if squad['player_count'] == 0:
                squad_impact      = 0.0
                squad_impact_norm = 50.0
            else:
                squad_impact      = round(role_fitness - squad_avg, 2)
                squad_impact_norm = round(min(100.0, max(0.0, 50.0 + squad_impact)), 2)

            combined = round(
                0.40 * role_fitness + 0.30 * squad_impact_norm + 0.30 * league_adaptation, 2
            )
            if combined >= 72:   verdict = 'Starter Upgrade'
            elif combined >= 58: verdict = 'Strong Signing'
            elif combined >= 44: verdict = 'Squad Depth'
            elif combined >= 32: verdict = 'Development Option'
            else:                verdict = 'Below Squad Level'

            # Try to get the human-readable team name from the Club JSON
            club_display = club_entry.name.replace('_', ' ')
            club_json_f  = next(
                (e for e in os.scandir(club_entry.path)
                 if e.name.startswith('Club_') and e.name.endswith('.json')),
                None
            )
            if club_json_f:
                try:
                    with open(club_json_f.path, encoding='utf-8') as f:
                        cd = _json.load(f)
                    club_display = cd.get('team_name', club_display)
                except Exception:
                    pass

            clubs.append({
                'club':              club_display,
                'club_slug':         club_entry.name,
                'club_path':         f"{tc}/{tcomp}/{club_entry.name}",
                'overall_score':     combined,
                'role_fitness':      round(role_fitness, 2),
                'squad_impact':      squad_impact,
                'league_adaptation': league_adaptation,
                'player_count':      squad['player_count'],
                'verdict':           verdict,
                'pool_info': {
                    'position_line': cand_line,
                    'pool_size':     squad['player_count'],
                    'small_sample':  squad['player_count'] < 5,
                },
            })

        clubs.sort(key=lambda x: x['overall_score'], reverse=True)
        return jsonify({
            'player_name':         cand['name'],
            'player_global_rank':  player_global_rank,
            'league':              tcomp.replace('_', ' '),
            'country':             tc.replace('_', ' '),
            'league_intensity':    round(target_intensity, 4),
            'target_global_rank':  target_rank,
            'move_label':          move_label,
            'clubs':               clubs,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/club_compatibility')
def api_club_compatibility():
    player_path = request.args.get('player', '')
    role_file   = request.args.get('role', '')
    club_path   = request.args.get('club', '')

    if not player_path or not role_file or not club_path:
        return jsonify({'error': 'player, role, and club are required'}), 400

    p_parts = player_path.split('/')
    c_parts = club_path.split('/')
    if len(p_parts) != 4:
        return jsonify({'error': 'invalid player path'}), 400
    if len(c_parts) != 3:
        return jsonify({'error': 'club path must be country/competition/club'}), 400

    player_fname = p_parts[3] if p_parts[3].endswith('.json') else p_parts[3] + '.json'
    player_full  = os.path.join(OUTPUT_DIR, p_parts[0], p_parts[1], p_parts[2], 'Players', player_fname)
    if not os.path.exists(player_full):
        return jsonify({'error': 'player file not found'}), 404

    try:
        from tactical_match_engine.services.json_loader import load_sofascore_player, get_league_intensity, get_league_info
        from tactical_match_engine.engine.role_encoder import get_role_fitness_vector
        from tactical_match_engine.engine.physical_adaptation import calculate_physical_adaptation

        # ── Candidate ────────────────────────────────────────────────────────
        cand        = load_sofascore_player(player_full)
        cand_result = get_role_fitness_vector(cand['raw_stats'], cand['position'], role_path=role_file, flat_weights=True)
        if cand_result is None:
            return jsonify({'error': 'no role profile available for this position'}), 404

        # ── Target club ───────────────────────────────────────────────────────
        tc, tcomp, tclub = c_parts
        t_info            = get_league_info(tc.replace('_', ' '), tcomp.replace('_', ' '))
        target_intensity  = t_info['intensity']
        target_rank       = t_info.get('global_rank')
        player_global_rank = cand.get('current_league_global_rank')
        age               = cand.get('age', 26)
        league_adaptation = round(
            calculate_physical_adaptation(cand['current_league_intensity'], target_intensity) * 100.0, 2
        )

        # Tier gap label
        if player_global_rank and target_rank:
            rank_diff = player_global_rank - target_rank
            if rank_diff > 5:
                move_label = f'+{rank_diff} ranks up'
            elif rank_diff < -5:
                move_label = f'{abs(rank_diff)} ranks down'
            else:
                move_label = 'lateral move'
        else:
            move_label = ''

        # ── Squad analysis ───────────────────────────────────────────────────
        cand_positions = [str(p).upper() for p in cand.get('positions', [cand['position']])]
        cand_line      = get_position_line(cand_positions)
        squad    = _get_squad_role_analysis(tc, tcomp, tclub, role_file,
                                            cand_line=cand_line, exclude_path=player_full)
        pool_info = {
            'position_line': cand_line,
            'pool_size':     squad['player_count'],
            'small_sample':  squad['player_count'] < 5,
        }

        # ── Scores ────────────────────────────────────────────────────────────
        role_fitness = cand_result['overall_score']
        squad_avg    = squad['avg_overall']
        if squad['player_count'] == 0:
            squad_impact      = 0.0
            squad_impact_norm = 50.0
        else:
            squad_impact      = round(role_fitness - squad_avg, 2)
            squad_impact_norm = round(min(100.0, max(0.0, 50.0 + squad_impact)), 2)

        combined = round(
            0.40 * role_fitness + 0.30 * squad_impact_norm + 0.30 * league_adaptation, 2
        )

        if combined >= 72:   verdict = 'Starter Upgrade'
        elif combined >= 58: verdict = 'Strong Signing'
        elif combined >= 44: verdict = 'Squad Depth'
        elif combined >= 32: verdict = 'Development Option'
        else:                verdict = 'Below Squad Level'

        # ── Metric deltas ─────────────────────────────────────────────────────
        sq_metric_avgs = squad['avg_metric_scores']
        sq_metric_raw  = squad['avg_metric_raw']
        metric_deltas  = [
            {
                'name':              m['name'],
                'category':          m['category'],
                'higher_is_better':  m['higher_is_better'],
                # candidate
                'candidate_raw':     m['raw_value'],
                'candidate_score':   m['score'],
                # squad avg (per-90 raw values averaged across position-matched players)
                'squad_avg_raw':     sq_metric_raw.get(m['name'], 0.0),
                'squad_avg_score':   sq_metric_avgs.get(m['name'], 0.0),
                # deltas
                'raw_delta':         round(m['raw_value']  - sq_metric_raw.get(m['name'], 0.0),  4),
                'score_delta':       round(m['score']      - sq_metric_avgs.get(m['name'], 0.0), 2),
            }
            for m in cand_result['metric_details']
        ]

        # ── Scouting narrative explanations ──────────────────────────────────
        from tactical_match_engine.engine.explanation_generator import generate_explanation
        from tactical_match_engine.engine.contender_simulation import simulate_contender_impact

        compat_scores_for_explain = {
            'tactical_similarity':  role_fitness / 100.0,
            'statistical_match':    min(1.0, squad_impact_norm / 100.0),
            'physical_adaptation':  league_adaptation / 100.0,
        }
        # Progression index for contender simulation
        _mins   = float(cand['raw_stats'].get('minutesPlayed', 90) or 90)
        _per90  = max(_mins / 90.0, 0.001)
        _ft     = float(cand['raw_stats'].get('accurateFinalThirdPasses', 0) or 0)
        _drb    = float(cand['raw_stats'].get('successfulDribbles', 0) or 0)
        _prog   = ((_ft / _per90) + (_drb / _per90)) / 2.0 if (_ft > 0 or _drb > 0) else 0.0
        _sq_ft  = float(squad.get('avg_metric_raw', {}).get('Final Third Passes', 0) or 0)
        _sq_drb = float(squad.get('avg_metric_raw', {}).get('Successful Dribbles', 0) or 0)
        _sq_prog = (_sq_ft + _sq_drb) / 2.0

        contender_sim = simulate_contender_impact(
            player_progression_index=round(_prog, 4),
            squad_average_progression=round(_sq_prog, 4),
        )
        explanations = generate_explanation(compat_scores_for_explain, contender_sim)

        return jsonify({
            'candidate': {
                'name':             cand['name'],
                'position':         cand['position'],
                'positions':        cand.get('positions', []),
                'age':              cand['age'],
                'league_intensity': round(cand['current_league_intensity'], 4),
                'global_rank':      cand.get('current_league_global_rank'),
                'league_name':      cand.get('current_league_name', ''),
                'role_fitness': {
                    'overall_score':   role_fitness,
                    'category_scores': cand_result['category_scores'],
                    'weights':         cand_result['weights'],
                },
            },
            'target': {
                'club':               tclub.replace('_', ' '),
                'competition':        tcomp.replace('_', ' '),
                'country':            tc.replace('_', ' '),
                'league_intensity':   round(target_intensity, 4),
                'global_rank':        target_rank,
                'squad_avg_score':    squad_avg,
                'avg_category_scores': squad['avg_category_scores'],
                'squad_players':      squad['players'],
                'player_count':       squad['player_count'],
            },
            'scores': {
                'role_fitness':      role_fitness,
                'squad_impact':      squad_impact,
                'squad_impact_norm': squad_impact_norm,
                'league_adaptation': league_adaptation,
                'combined_score':    combined,
            },
            'move_label':       move_label,
            'metric_deltas':    metric_deltas,
            'verdict':          verdict,
            'pool_info':        pool_info,
            'contender_simulation': contender_sim,
            'explanations':     explanations,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Compare players ───────────────────────────────────────────────────────────
@app.route('/compare')
def compare():
    return render_template('compare.html')

@app.route('/api/player_search')
def api_player_search():
    q      = request.args.get('q', '').lower()
    result = get_all_players_flat(OUTPUT_DIR, query=q, limit=20)
    return jsonify(result)

@app.route('/api/league_position_avg')
def api_league_position_avg():
    """
    Return league-wide average stats for players sharing the same position as
    the given player (excluding the player themselves).

    Query params:
      player   – "Country/Competition/Club/filename.json"
      pos_mode – 'exact' (default) | 'grouped'
    """
    player_path = request.args.get('player', '').strip()
    pos_mode    = request.args.get('pos_mode', 'exact')
    if pos_mode not in ('exact', 'grouped'):
        pos_mode = 'exact'
    if not player_path:
        return jsonify({'error': 'player path required'}), 400

    from data_loader import get_league_position_avg
    result = get_league_position_avg(OUTPUT_DIR, player_path, pos_mode)
    if result is None:
        return jsonify({'error': 'player not found or invalid path'}), 404
    if 'error' in result:
        return jsonify(result), 400
    return jsonify(result)

@app.route('/api/player_detail')
def api_player_detail():
    path = request.args.get('path', '')
    parts = path.split('/')
    if len(parts) == 4:
        data = get_player(OUTPUT_DIR, parts[0], parts[1], parts[2], parts[3])
        heatmap = get_heatmap_path(OUTPUT_DIR, parts[0], parts[1], parts[2], parts[3])
        return jsonify({'data': data, 'heatmap': heatmap})
    return jsonify({'error': 'invalid path'}), 400

# ── Export ────────────────────────────────────────────────────────────────────
@app.route('/export')
def export_page():
    competitions = get_all_competitions(OUTPUT_DIR)
    return render_template('export.html', competitions=competitions)

@app.route('/api/export/csv')
def export_csv():
    filters = {k: v for k, v in request.args.items()}
    filters['page']     = 1
    filters['per_page'] = 99999
    result   = search_players(OUTPUT_DIR, filters)
    players  = result.get('players', [])

    si     = io.StringIO()
    writer = csv.DictWriter(si, fieldnames=[
        'name','team','competition','country','position','nationality',
        'age','height','preferred_foot','shirt_number',
        'appearances','goals','assists','rating',
        'pass_accuracy','key_passes','accurate_long_balls',
        'dribbles','duels_won','aerial_won',
        'tackles','interceptions','clearances',
        'shots_on_target','big_chances_missed',
        'yellow_cards','red_cards','minutes_played'
    ])
    writer.writeheader()
    for p in players:
        s = p.get('stats', {})
        writer.writerow({
            'name':               p.get('name',''),
            'team':               p.get('team',''),
            'competition':        p.get('competition',''),
            'country':            p.get('country',''),
            'position':           p.get('position',''),
            'nationality':        p.get('nationality',''),
            'age':                p.get('age',''),
            'height':             p.get('height',''),
            'preferred_foot':     p.get('preferred_foot',''),
            'shirt_number':       p.get('shirt_number',''),
            'appearances':        s.get('appearances',''),
            'goals':              s.get('goals',''),
            'assists':            s.get('assists',''),
            'rating':             round(s.get('rating',0),2) if s.get('rating') else '',
            'pass_accuracy':      s.get('accuratePassesPercentage',''),
            'key_passes':         s.get('keyPasses',''),
            'accurate_long_balls':s.get('accurateLongBalls',''),
            'dribbles':           s.get('successfulDribbles',''),
            'duels_won':          s.get('totalDuelsWon',''),
            'aerial_won':         s.get('aerialDuelsWon',''),
            'tackles':            s.get('tackles',''),
            'interceptions':      s.get('interceptions',''),
            'clearances':         s.get('clearances',''),
            'shots_on_target':    s.get('shotsOnTarget',''),
            'big_chances_missed': s.get('bigChancesMissed',''),
            'yellow_cards':       s.get('yellowCards',''),
            'red_cards':          s.get('redCards',''),
            'minutes_played':     s.get('minutesPlayed',''),
        })

    output = si.getvalue()
    return send_file(
        io.BytesIO(output.encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name='players_export.csv'
    )

# ── Scrape jobs ──────────────────────────────────────────────────────────────
_scrape_jobs: dict = {}  # job_id -> {lines, progress, done, error}


def _run_scrape_job(job_id: str, tid: str, uniq_tid: str,
                    season_id: str, skip: bool) -> None:
    """Run run_scrape() directly in this thread, just like ui/tab_league.py does."""
    job = _scrape_jobs[job_id]

    def _cb(line: str) -> None:
        """Callback wired into run_scrape() instead of print()."""
        if line.startswith('PROGRESS:'):
            try:
                job['progress'] = int(line[9:])
            except ValueError:
                pass
        job['lines'].append(line)

    try:
        from scraper_cli import run_scrape
        run_scrape(
            tid=tid,
            uniq_tid=uniq_tid,
            season_id=season_id,
            out_dir=OUTPUT_DIR,
            skip_existing=skip,
            log_fn=_cb,
        )
        job['done'] = True
    except Exception as e:
        job['lines'].append(f'ERROR: {e}')
        job['done'] = True
        job['error'] = True


@app.route('/api/scrape', methods=['POST'])
def api_scrape_start():
    body      = request.get_json(silent=True) or {}
    tid       = str(body.get('tid', '')).strip()
    uniq_tid  = str(body.get('uniq_tid', '')).strip()
    season_id = str(body.get('season_id', '')).strip()
    skip      = bool(body.get('skip', False))

    if not tid or not uniq_tid or not season_id:
        return jsonify({'error': 'tid, uniq_tid and season_id are required'}), 400

    job_id = str(uuid.uuid4())
    _scrape_jobs[job_id] = {
        'lines': [], 'progress': 0, 'done': False, 'error': False
    }

    threading.Thread(
        target=_run_scrape_job,
        args=(job_id, tid, uniq_tid, season_id, skip),
        daemon=True,
    ).start()

    return jsonify({'job_id': job_id})


@app.route('/api/scrape/<job_id>')
def api_scrape_status(job_id):
    job = _scrape_jobs.get(job_id)
    if job is None:
        return jsonify({'error': 'unknown job'}), 404
    offset = int(request.args.get('offset', 0))
    return jsonify({
        'done':     job['done'],
        'error':    job['error'],
        'progress': job['progress'],
        'lines':    job['lines'][offset:],
        'total':    len(job['lines']),
    })


# ── IFFHS League Rankings ────────────────────────────────────────────────────
@app.route('/league_rankings')
def league_rankings():
    return render_template('league_rankings.html')

@app.route('/api/league_rankings')
def api_league_rankings():
    """Return the full Opta+IFFHS rankings list, optionally filtered by confederation."""
    conf_filter = request.args.get('confederation', '').strip().lower()
    try:
        from tactical_match_engine.services.json_loader import load_iffhs_data
        data = load_iffhs_data()
        entries = []
        for entry in data.get('rankings', []):
            conf = entry.get('confederation', '').lower()
            if conf_filter and conf != conf_filter:
                continue
            entries.append({
                'rank':          entry.get('rank'),
                'country':       entry.get('country'),
                'confederation': entry.get('confederation'),
                'iffhs_points':  entry.get('iffhs_points'),
                'iffhs_rank':    entry.get('iffhs_rank'),
                's_country':     round(entry.get('s_country', 0), 4),
                'retention':     round(entry.get('retention', 0), 4),
                'divisions':     entry.get('divisions', []),
            })
        return jsonify({
            'source':                  data.get('source'),
            'endpoint':                data.get('endpoint'),
            'iffhs_source':            data.get('iffhs_source'),
            'iffhs_matched_countries': data.get('iffhs_matched_countries'),
            'parameters':              data.get('parameters', {}),
            'rankings':                entries,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Player percentiles within their league ────────────────────────────────────
@app.route('/api/player_percentiles')
def api_player_percentiles():
    """
    Compute percentile rank for every numeric stat of a player
    compared to all players of the same position line in their competition.
    """
    player_path = request.args.get('player', '').strip()
    if not player_path:
        return jsonify({'error': 'player is required'}), 400
    parts = player_path.split('/')
    if len(parts) != 4:
        return jsonify({'error': 'invalid player path'}), 400

    country, competition, club_slug, filename = parts
    target_file = filename if filename.endswith('.json') else filename + '.json'
    full_path = os.path.join(OUTPUT_DIR, country, competition, club_slug, 'Players', target_file)
    if not os.path.exists(full_path):
        return jsonify({'error': 'player file not found'}), 404

    try:
        import glob as _glob
        import json as _json
        from tactical_match_engine.engine.normalization import percentile_score

        with open(full_path, encoding='utf-8') as f:
            raw = _json.load(f)
        player_stats = raw.get('statistics', {}).get('statistics', {})
        player_pos_raw = raw.get('positions') or [raw.get('profile', {}).get('player', {}).get('position', '')]
        player_line = get_position_line([str(p).upper() for p in player_pos_raw])

        # Collect pool: all players in same competition with same position line
        pool_stats: dict[str, list] = {}
        comp_dir = os.path.join(OUTPUT_DIR, country, competition)
        for fpath in _glob.glob(os.path.join(comp_dir, '*', 'Players', '*.json')):
            try:
                with open(fpath, encoding='utf-8') as f:
                    pd = _json.load(f)
                pos_raw = pd.get('positions') or [pd.get('profile', {}).get('player', {}).get('position', '')]
                if get_position_line([str(p).upper() for p in pos_raw]) != player_line:
                    continue
                ps = pd.get('statistics', {}).get('statistics', {})
                for k, v in ps.items():
                    if isinstance(v, (int, float)) and not isinstance(v, bool) and v is not None:
                        pool_stats.setdefault(k, []).append(float(v))
            except Exception:
                continue

        # Stats where lower is better
        LOWER_IS_BETTER = {'yellowCards', 'redCards', 'fouls', 'ownGoals', 'bigChancesMissed'}

        percentiles = {}
        for stat, val in player_stats.items():
            if not isinstance(val, (int, float)) or isinstance(val, bool) or val is None:
                continue
            pop = pool_stats.get(stat, [])
            if len(pop) < 2:
                continue
            hib = stat not in LOWER_IS_BETTER
            pct = percentile_score(float(val), pop, higher_is_better=hib)
            percentiles[stat] = round(pct, 1)

        return jsonify({
            'player_path':  player_path,
            'position_line': player_line,
            'pool_size':    len(pool_stats.get('appearances', [])),
            'competition':  competition.replace('_', ' '),
            'percentiles':  percentiles,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Export Club Fit analysis as CSV ──────────────────────────────────────────
@app.route('/api/export_club_fit_csv')
def api_export_club_fit_csv():
    """
    Run /api/club_compatibility and return the result as a downloadable CSV
    containing the full metric deltas table + summary scores + explanations.
    """
    player_path = request.args.get('player', '')
    role_file   = request.args.get('role', '')
    club_path   = request.args.get('club', '')

    if not player_path or not role_file or not club_path:
        return jsonify({'error': 'player, role, and club are required'}), 400

    # Re-use the existing logic by doing an internal fetch via the service layer
    p_parts = player_path.split('/')
    c_parts = club_path.split('/')
    if len(p_parts) != 4 or len(c_parts) != 3:
        return jsonify({'error': 'invalid paths'}), 400

    player_fname = p_parts[3] if p_parts[3].endswith('.json') else p_parts[3] + '.json'
    player_full  = os.path.join(OUTPUT_DIR, p_parts[0], p_parts[1], p_parts[2], 'Players', player_fname)
    if not os.path.exists(player_full):
        return jsonify({'error': 'player file not found'}), 404

    try:
        from tactical_match_engine.services.json_loader import load_sofascore_player, get_league_intensity
        from tactical_match_engine.engine.role_encoder import get_role_fitness_vector
        from tactical_match_engine.engine.physical_adaptation import calculate_physical_adaptation
        from tactical_match_engine.engine.explanation_generator import generate_explanation
        from tactical_match_engine.engine.contender_simulation import simulate_contender_impact

        cand        = load_sofascore_player(player_full)
        cand_result = get_role_fitness_vector(cand['raw_stats'], cand['position'], role_path=role_file, flat_weights=True)
        if cand_result is None:
            return jsonify({'error': 'no role profile available for this position'}), 404

        tc, tcomp, tclub = c_parts
        target_intensity  = get_league_intensity(tc.replace('_', ' '), tcomp.replace('_', ' '))
        league_adaptation = round(calculate_physical_adaptation(cand['current_league_intensity'], target_intensity) * 100.0, 2)

        cand_positions = [str(p).upper() for p in cand.get('positions', [cand['position']])]
        cand_line      = get_position_line(cand_positions)
        squad    = _get_squad_role_analysis(tc, tcomp, tclub, role_file, cand_line=cand_line, exclude_path=player_full)

        role_fitness     = cand_result['overall_score']
        squad_avg        = squad['avg_overall']
        squad_impact     = round(role_fitness - squad_avg, 2) if squad['player_count'] else 0.0
        squad_impact_norm = round(min(100.0, max(0.0, 50.0 + squad_impact)), 2) if squad['player_count'] else 50.0
        combined = round(0.40 * role_fitness + 0.30 * squad_impact_norm + 0.30 * league_adaptation, 2)
        if combined >= 72:   verdict = 'Starter Upgrade'
        elif combined >= 58: verdict = 'Strong Signing'
        elif combined >= 44: verdict = 'Squad Depth'
        elif combined >= 32: verdict = 'Development Option'
        else:                verdict = 'Below Squad Level'

        sq_metric_avgs = squad['avg_metric_scores']
        sq_metric_raw  = squad['avg_metric_raw']
        metric_deltas  = [
            {
                'name':            m['name'],
                'category':        m['category'],
                'higher_is_better': m['higher_is_better'],
                'candidate_raw':   m['raw_value'],
                'candidate_score': m['score'],
                'squad_avg_raw':   sq_metric_raw.get(m['name'], 0.0),
                'squad_avg_score': sq_metric_avgs.get(m['name'], 0.0),
                'raw_delta':       round(m['raw_value'] - sq_metric_raw.get(m['name'], 0.0), 4),
                'score_delta':     round(m['score']     - sq_metric_avgs.get(m['name'], 0.0), 2),
            }
            for m in cand_result['metric_details']
        ]

        compat_scores = {
            'tactical_similarity': role_fitness / 100.0,
            'statistical_match':   min(1.0, squad_impact_norm / 100.0),
            'physical_adaptation': league_adaptation / 100.0,
        }
        _mins  = float(cand['raw_stats'].get('minutesPlayed', 90) or 90)
        _per90 = max(_mins / 90.0, 0.001)
        _ft    = float(cand['raw_stats'].get('accurateFinalThirdPasses', 0) or 0)
        _drb   = float(cand['raw_stats'].get('successfulDribbles', 0) or 0)
        _prog  = ((_ft / _per90) + (_drb / _per90)) / 2.0 if (_ft > 0 or _drb > 0) else 0.0
        _sq_prog = (float(squad['avg_metric_raw'].get('Final Third Passes', 0) or 0) +
                    float(squad['avg_metric_raw'].get('Successful Dribbles', 0) or 0)) / 2.0
        contender_sim = simulate_contender_impact(round(_prog, 4), round(_sq_prog, 4))
        explanations  = generate_explanation(compat_scores, contender_sim)

        output = io.StringIO()
        writer = csv.writer(output)

        # Section 1: Summary
        writer.writerow(['SCOUTING REPORT — CLUB FIT ANALYSIS'])
        writer.writerow([])
        writer.writerow(['Player', cand['name']])
        writer.writerow(['Position', '/'.join(cand.get('positions', [cand['position']]))])
        writer.writerow(['Age', cand['age']])
        writer.writerow(['Current League Intensity', f"{cand['current_league_intensity'] * 100:.1f}%"])
        writer.writerow(['Role', role_file.split('/')[-1].replace('.json', '').replace('_', ' ')])
        writer.writerow(['Target Club', tclub.replace('_', ' ')])
        writer.writerow(['Target League', tcomp.replace('_', ' ')])
        writer.writerow(['Target Country', tc.replace('_', ' ')])
        writer.writerow(['Target League Intensity', f"{target_intensity * 100:.1f}%"])
        writer.writerow([])

        writer.writerow(['SCORES'])
        writer.writerow(['Combined Score', combined])
        writer.writerow(['Verdict', verdict])
        writer.writerow(['Role Fitness', role_fitness])
        writer.writerow(['Squad Impact', squad_impact])
        writer.writerow(['League Adaptation', league_adaptation])
        writer.writerow([])

        writer.writerow(['CONTENDER PROJECTION'])
        writer.writerow(['Progression Gain', contender_sim['progression_gain']])
        writer.writerow(['xG Gain per Match', contender_sim['xg_gain_per_match']])
        writer.writerow(['Season xG Gain', contender_sim['season_xg_gain']])
        writer.writerow(['Projected Goal Gain', contender_sim['goal_gain']])
        writer.writerow(['Points Gain', contender_sim['points_gain']])
        writer.writerow(['Title Probability Shift', f"{contender_sim['title_probability_shift']:+.4f}"])
        writer.writerow([])

        writer.writerow(['SCOUTING NARRATIVE'])
        writer.writerow(['Why Club Needs Player', explanations.get('why_club_needs_player', '')])
        writer.writerow(['Why Player Fits Club', explanations.get('why_player_fits_club', '')])
        writer.writerow(['Why Club Becomes Contender', explanations.get('why_club_becomes_contender', '')])
        writer.writerow(['Risk Assessment', explanations.get('risk_assessment', '')])
        writer.writerow([])

        writer.writerow(['METRIC DETAILS'])
        writer.writerow(['Metric', 'Category', 'Higher Is Better', 'Candidate /90', 'Squad Avg /90',
                         'Raw Delta', 'Candidate Score', 'Squad Avg Score', 'Score Delta'])
        for m in metric_deltas:
            writer.writerow([
                m['name'], m['category'], m['higher_is_better'],
                m['candidate_raw'], m['squad_avg_raw'], m['raw_delta'],
                m['candidate_score'], m['squad_avg_score'], m['score_delta'],
            ])

        output.seek(0)
        safe_player = cand['name'].replace(' ', '_')
        safe_club   = tclub.replace('_', ' ').replace(' ', '_')
        filename_csv = f"ClubFit_{safe_player}_{safe_club}.csv"
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename_csv,
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Export Club Fit analysis as PDF ──────────────────────────────────────────
@app.route('/api/export_club_fit_pdf', methods=['POST'])
def api_export_club_fit_pdf():
    """
    Run the Club Fit engine for the given player/role/club combination
    and return the result as a downloadable PDF scouting report.
    """
    body       = request.get_json(silent=True) or {}
    player_path = body.get('player', '')
    role_file   = body.get('role', '')
    club_path   = body.get('club', '')

    if not player_path or not role_file or not club_path:
        return jsonify({'error': 'player, role, and club are required'}), 400

    p_parts = player_path.split('/')
    c_parts = club_path.split('/')
    if len(p_parts) != 4 or len(c_parts) != 3:
        return jsonify({'error': 'invalid paths'}), 400

    player_fname = p_parts[3] if p_parts[3].endswith('.json') else p_parts[3] + '.json'
    player_full  = os.path.join(OUTPUT_DIR, p_parts[0], p_parts[1], p_parts[2], 'Players', player_fname)
    if not os.path.exists(player_full):
        return jsonify({'error': 'player file not found'}), 404

    try:
        from tactical_match_engine.services.json_loader import (
            load_sofascore_player, get_league_intensity, get_league_info,
        )
        from tactical_match_engine.engine.role_encoder import get_role_fitness_vector
        from tactical_match_engine.engine.physical_adaptation import calculate_physical_adaptation
        from tactical_match_engine.engine.explanation_generator import generate_explanation
        from tactical_match_engine.engine.contender_simulation import simulate_contender_impact
        from tactical_match_engine.services.pdf_generator import generate_pdf

        cand        = load_sofascore_player(player_full)
        cand_result = get_role_fitness_vector(
            cand['raw_stats'], cand['position'], role_path=role_file, flat_weights=True
        )
        if cand_result is None:
            return jsonify({'error': 'no role profile available for this position'}), 404

        tc, tcomp, tclub = c_parts
        t_info           = get_league_info(tc.replace('_', ' '), tcomp.replace('_', ' '))
        target_intensity  = t_info['intensity']
        target_rank       = t_info.get('global_rank')
        player_global_rank = cand.get('current_league_global_rank')
        league_adaptation = round(
            calculate_physical_adaptation(cand['current_league_intensity'], target_intensity) * 100.0, 2
        )

        if player_global_rank and target_rank:
            rank_diff = player_global_rank - target_rank
            if rank_diff > 5:    move_label = f'+{rank_diff} ranks up'
            elif rank_diff < -5: move_label = f'{abs(rank_diff)} ranks down'
            else:                move_label = 'lateral move'
        else:
            move_label = ''

        cand_positions = [str(p).upper() for p in cand.get('positions', [cand['position']])]
        cand_line      = get_position_line(cand_positions)
        squad    = _get_squad_role_analysis(tc, tcomp, tclub, role_file,
                                            cand_line=cand_line, exclude_path=player_full)

        role_fitness      = cand_result['overall_score']
        squad_avg         = squad['avg_overall']
        squad_impact      = round(role_fitness - squad_avg, 2) if squad['player_count'] else 0.0
        squad_impact_norm = round(min(100.0, max(0.0, 50.0 + squad_impact)), 2) if squad['player_count'] else 50.0
        combined = round(0.40 * role_fitness + 0.30 * squad_impact_norm + 0.30 * league_adaptation, 2)
        if combined >= 72:   verdict = 'Starter Upgrade'
        elif combined >= 58: verdict = 'Strong Signing'
        elif combined >= 44: verdict = 'Squad Depth'
        elif combined >= 32: verdict = 'Development Option'
        else:                verdict = 'Below Squad Level'

        sq_metric_avgs = squad['avg_metric_scores']
        sq_metric_raw  = squad['avg_metric_raw']
        metric_deltas  = [
            {
                'name':             m['name'],
                'category':         m['category'],
                'higher_is_better': m['higher_is_better'],
                'candidate_raw':    m['raw_value'],
                'candidate_score':  m['score'],
                'squad_avg_raw':    sq_metric_raw.get(m['name'], 0.0),
                'squad_avg_score':  sq_metric_avgs.get(m['name'], 0.0),
                'raw_delta':        round(m['raw_value'] - sq_metric_raw.get(m['name'], 0.0), 4),
                'score_delta':      round(m['score']     - sq_metric_avgs.get(m['name'], 0.0), 2),
            }
            for m in cand_result['metric_details']
        ]

        compat_scores = {
            'tactical_similarity': role_fitness / 100.0,
            'statistical_match':   min(1.0, squad_impact_norm / 100.0),
            'physical_adaptation': league_adaptation / 100.0,
        }
        _mins   = float(cand['raw_stats'].get('minutesPlayed', 90) or 90)
        _per90  = max(_mins / 90.0, 0.001)
        _ft     = float(cand['raw_stats'].get('accurateFinalThirdPasses', 0) or 0)
        _drb    = float(cand['raw_stats'].get('successfulDribbles', 0) or 0)
        _prog   = ((_ft / _per90) + (_drb / _per90)) / 2.0 if (_ft > 0 or _drb > 0) else 0.0
        _sq_prog = (float(squad['avg_metric_raw'].get('Final Third Passes', 0) or 0) +
                    float(squad['avg_metric_raw'].get('Successful Dribbles', 0) or 0)) / 2.0
        contender_sim = simulate_contender_impact(round(_prog, 4), round(_sq_prog, 4))
        explanations  = generate_explanation(compat_scores, contender_sim)

        result_dict = {
            'candidate': {
                'name':             cand['name'],
                'position':         cand['position'],
                'positions':        cand.get('positions', []),
                'age':              cand['age'],
                'league_intensity': round(cand['current_league_intensity'], 4),
                'global_rank':      player_global_rank,
                'league_name':      cand.get('current_league_name', ''),
                'role_fitness': {
                    'overall_score':   role_fitness,
                    'category_scores': cand_result['category_scores'],
                },
            },
            'target': {
                'club':        tclub.replace('_', ' '),
                'competition': tcomp.replace('_', ' '),
                'country':     tc.replace('_', ' '),
                'squad_avg_score': squad_avg,
            },
            'scores': {
                'role_fitness':      role_fitness,
                'squad_impact':      squad_impact,
                'squad_impact_norm': squad_impact_norm,
                'league_adaptation': league_adaptation,
                'combined_score':    combined,
            },
            'verdict':              verdict,
            'move_label':           move_label,
            'metric_deltas':        metric_deltas,
            'contender_simulation': contender_sim,
            'explanations':         explanations,
        }

        pdf_bytes = generate_pdf(result_dict)
        safe_player = cand['name'].replace(' ', '_')
        safe_club   = tclub.replace('_', ' ').replace(' ', '_')
        filename_pdf = f"ClubFit_{safe_player}_{safe_club}.pdf"
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename_pdf,
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Batch Comparison ──────────────────────────────────────────────────────────
@app.route('/batch_compare')
def batch_compare():
    competitions = get_all_competitions(OUTPUT_DIR)
    return render_template('batch_compare.html', competitions=competitions)


@app.route('/api/batch_compare', methods=['POST'])
def api_batch_compare():
    """
    Score a list of players against a single target club + role.
    Body JSON::
        {
            "players": ["Country/Competition/Club/file.json", ...],  # max 30
            "role":    "2. Mid-Line/I_Central_Midfielder.json",
            "club":    "Country/Competition/Club"
        }
    Returns a ranked list of per-player results.
    """
    import concurrent.futures

    body        = request.get_json(silent=True) or {}
    player_list = body.get('players', [])
    role_file   = str(body.get('role', '')).strip()
    club_path   = str(body.get('club', '')).strip()

    if not player_list or not role_file or not club_path:
        return jsonify({'error': 'players, role, and club are required'}), 400
    if len(player_list) > 30:
        return jsonify({'error': 'maximum 30 players per batch'}), 400

    c_parts = club_path.split('/')
    if len(c_parts) != 3:
        return jsonify({'error': 'club path must be Country/Competition/Club'}), 400

    try:
        from tactical_match_engine.services.json_loader import (
            load_sofascore_player, get_league_info,
        )
        from tactical_match_engine.engine.role_encoder import get_role_fitness_vector
        from tactical_match_engine.engine.physical_adaptation import calculate_physical_adaptation

        tc, tcomp, tclub = c_parts
        t_info           = get_league_info(tc.replace('_', ' '), tcomp.replace('_', ' '))
        target_intensity = t_info['intensity']

        # Compute squad stats once (shared across all players)
        squad = _get_squad_role_analysis(tc, tcomp, tclub, role_file, cand_line=None)
        squad_avg = squad['avg_overall']

        def _score_player(player_path: str) -> dict:
            p_parts  = player_path.strip().split('/')
            if len(p_parts) != 4:
                return {'error': 'invalid path', 'player_path': player_path}
            player_fname = p_parts[3] if p_parts[3].endswith('.json') else p_parts[3] + '.json'
            player_full  = os.path.join(
                OUTPUT_DIR, p_parts[0], p_parts[1], p_parts[2], 'Players', player_fname
            )
            if not os.path.exists(player_full):
                return {'error': 'file not found', 'player_path': player_path}
            try:
                cand        = load_sofascore_player(player_full)
                cand_result = get_role_fitness_vector(
                    cand['raw_stats'], cand['position'],
                    role_path=role_file, flat_weights=True,
                )
                if cand_result is None:
                    return {'error': 'no role profile', 'player_path': player_path, 'name': cand.get('name', '')}

                role_fitness      = cand_result['overall_score']
                league_adaptation = round(
                    calculate_physical_adaptation(cand['current_league_intensity'], target_intensity) * 100.0, 2
                )
                squad_impact      = round(role_fitness - squad_avg, 2) if squad['player_count'] else 0.0
                squad_impact_norm = round(min(100.0, max(0.0, 50.0 + squad_impact)), 2) if squad['player_count'] else 50.0
                combined = round(0.40 * role_fitness + 0.30 * squad_impact_norm + 0.30 * league_adaptation, 2)

                if combined >= 72:   verdict = 'Starter Upgrade'
                elif combined >= 58: verdict = 'Strong Signing'
                elif combined >= 44: verdict = 'Squad Depth'
                elif combined >= 32: verdict = 'Development Option'
                else:                verdict = 'Below Squad Level'

                return {
                    'player_path':       player_path,
                    'name':              cand['name'],
                    'position':          cand['position'],
                    'age':               cand.get('age', ''),
                    'team':              cand.get('stats', {}).get('team', p_parts[2].replace('_', ' ')),
                    'league':            cand.get('current_league_name', ''),
                    'league_intensity':  round(cand['current_league_intensity'], 4),
                    'combined_score':    combined,
                    'role_fitness':      round(role_fitness, 2),
                    'squad_impact_norm': squad_impact_norm,
                    'league_adaptation': league_adaptation,
                    'verdict':           verdict,
                }
            except Exception as exc:
                return {'error': str(exc), 'player_path': player_path}

        with concurrent.futures.ThreadPoolExecutor(max_workers=6) as pool:
            raw_results = list(pool.map(_score_player, player_list))

        ok      = [r for r in raw_results if 'error' not in r]
        errors  = [r for r in raw_results if 'error' in r]
        ok.sort(key=lambda x: x['combined_score'], reverse=True)
        for i, r in enumerate(ok, 1):
            r['rank'] = i

        return jsonify({
            'club':         tclub.replace('_', ' '),
            'competition':  tcomp.replace('_', ' '),
            'country':      tc.replace('_', ' '),
            'role':         role_file.split('/')[-1].replace('.json', '').replace('_', ' '),
            'squad_avg':    round(squad_avg, 2),
            'results':      ok,
            'errors':       errors,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export_shortlist_xlsx', methods=['POST'])
def api_export_shortlist_xlsx():
    """
    Convert a batch comparison result set to an XLSX file for download.
    Body JSON::
        {
            "results":   [...],    # list of scored player dicts from /api/batch_compare
            "club_name": "FC ...",
            "role_name": "Deep-Lying Playmaker"
        }
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    body      = request.get_json(silent=True) or {}
    results   = body.get('results', [])
    club_name = str(body.get('club_name', 'Club')).strip()
    role_name = str(body.get('role_name', 'Role')).strip()

    if not results:
        return jsonify({'error': 'results are required'}), 400

    # ── Colour helpers ────────────────────────────────────────────────────────
    def _verdict_fill(verdict: str) -> PatternFill:
        mapping = {
            'Starter Upgrade':   '22c55e',
            'Strong Signing':    '0ea5e9',
            'Squad Depth':       'f59e0b',
            'Development Option':'a855f7',
            'Below Squad Level': 'ef4444',
        }
        return PatternFill('solid', fgColor=mapping.get(verdict, '64748b'))

    def _score_fill(score: float) -> PatternFill:
        if score >= 72:  return PatternFill('solid', fgColor='22c55e')
        if score >= 44:  return PatternFill('solid', fgColor='f59e0b')
        return PatternFill('solid', fgColor='ef4444')

    thin = Border(
        left=Side(style='thin', color='1e293b'),
        right=Side(style='thin', color='1e293b'),
        top=Side(style='thin', color='1e293b'),
        bottom=Side(style='thin', color='1e293b'),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Shortlist'

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells('A1:J1')
    ws['A1'] = f'Shortlist — {club_name} | Role: {role_name}'
    ws['A1'].font      = Font(bold=True, size=13, color='E2E8F0')
    ws['A1'].fill      = PatternFill('solid', fgColor='0d1b2a')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        'Rank', 'Name', 'Position', 'Age', 'Team', 'League',
        'Combined', 'Role Fit', 'Squad Impact', 'League Adapt', 'Verdict',
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font      = Font(bold=True, color='E2E8F0', size=9)
        cell.fill      = PatternFill('solid', fgColor='1e293b')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin
    ws.row_dimensions[2].height = 16

    # ── Data rows ─────────────────────────────────────────────────────────────
    for r in results:
        row_num = ws.max_row + 1
        combined = float(r.get('combined_score', 0))
        verdict  = r.get('verdict', '')
        row_data = [
            r.get('rank', ''),
            r.get('name', ''),
            r.get('position', ''),
            r.get('age', ''),
            r.get('team', ''),
            r.get('league', ''),
            combined,
            float(r.get('role_fitness', 0)),
            float(r.get('squad_impact_norm', 0)),
            float(r.get('league_adaptation', 0)),
            verdict,
        ]
        ws.append(row_data)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = thin
            cell.font      = Font(size=9, color='E2E8F0')
            # Dark alternating row bg
            bg = '0f172a' if row_num % 2 == 0 else '0d1117'
            cell.fill = PatternFill('solid', fgColor=bg)
        # Score colours for Combined column (col 7)
        ws.cell(row=row_num, column=7).fill  = _score_fill(combined)
        ws.cell(row=row_num, column=7).font  = Font(bold=True, size=9, color='0d1117')
        ws.cell(row=row_num, column=11).fill = _verdict_fill(verdict)
        ws.cell(row=row_num, column=11).font = Font(bold=True, size=9, color='0d1117')
        ws.row_dimensions[row_num].height = 15

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [6, 24, 10, 6, 22, 22, 10, 10, 13, 13, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_club = club_name.replace(' ', '_')
    safe_role = role_name.replace(' ', '_')
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Shortlist_{safe_club}_{safe_role}.xlsx',
    )


# ── Scatter ───────────────────────────────────────────────────────────────────────
CLUB_STAT_FIELDS = [
    'goalsScored', 'goalsConceded', 'assists', 'shots', 'penaltyGoals',
    'penaltiesTaken', 'successfulDribbles', 'dribbleAttempts', 'corners',
    'averageBallPossession', 'totalPasses', 'accuratePasses',
    'accuratePassesPercentage', 'totalLongBalls', 'accurateLongBalls',
    'accurateLongBallsPercentage', 'totalCrosses', 'accurateCrosses',
    'accurateCrossesPercentage', 'cleanSheets', 'interceptions', 'saves',
    'errorsLeadingToShot', 'totalDuels', 'duelsWon', 'duelsWonPercentage',
    'totalAerialDuels', 'aerialDuelsWon', 'aerialDuelsWonPercentage',
    'offsides', 'fouls', 'yellowCards', 'yellowRedCards', 'redCards',
    'shotsAgainst', 'goalKicks', 'ballRecovery', 'freeKicks', 'matches',
]

# Club stats that are percentages / ratios — never compute p90
_CLUB_P90_EXCLUDE = frozenset([
    'accuratePassesPercentage', 'accurateLongBallsPercentage',
    'accurateCrossesPercentage', 'duelsWonPercentage',
    'aerialDuelsWonPercentage', 'averageBallPossession',
])

@app.route('/scatter')
def scatter():
    competitions = get_all_competitions(OUTPUT_DIR)
    return render_template('scatter.html', competitions=competitions)

@app.route('/api/scatter_data')
def api_scatter_data():
    import glob as _glob
    import json as _json

    competition = request.args.get('competition', '').strip()
    country     = request.args.get('country',     '').strip()
    if not competition or not country:
        return jsonify({'error': 'competition and country are required'}), 400

    league_dir = os.path.join(OUTPUT_DIR, country, competition)
    if not os.path.isdir(league_dir):
        return jsonify({'error': 'competition directory not found'}), 404

    clubs     = []
    keys_seen = []   # preserves CLUB_STAT_FIELDS order, only adds keys once

    for club_entry in sorted(os.scandir(league_dir), key=lambda e: e.name):
        if not club_entry.is_dir():
            continue
        matches = _glob.glob(os.path.join(club_entry.path, 'Club_*_Season_*.json'))
        if not matches:
            continue
        try:
            with open(matches[0], encoding='utf-8') as f:
                raw = _json.load(f)
            stat_raw = raw.get('season_statistics', {}).get('statistics', {})
            stats     = {}
            stats_p90 = {}
            num_matches = stat_raw.get('matches')
            for field in CLUB_STAT_FIELDS:
                val = stat_raw.get(field)
                if val is not None and isinstance(val, (int, float)) and not isinstance(val, bool):
                    stats[field] = val
                    if field not in keys_seen:
                        keys_seen.append(field)
                    # Compute p90 for countable stats
                    if field not in _CLUB_P90_EXCLUDE and isinstance(num_matches, (int, float)) and num_matches > 0:
                        stats_p90[field] = round(val / num_matches, 2)
            # Derive display name from JSON if available, else slug
            club_name = raw.get('team_name', club_entry.name.replace('_', ' '))
            clubs.append({
                'club':      club_name,
                'club_slug': club_entry.name,
                'matches':   int(num_matches) if isinstance(num_matches, (int, float)) else 0,
                'stats':     stats,
                'stats_p90': stats_p90,
            })
        except Exception:
            continue

    # stat_keys = union in CLUB_STAT_FIELDS order
    stat_keys = [k for k in CLUB_STAT_FIELDS if k in keys_seen]
    return jsonify({'clubs': clubs, 'stat_keys': stat_keys})


# ── Player Scatter ────────────────────────────────────────────────────────────
PLAYER_STAT_FIELDS = [
    'goals', 'assists', 'shots', 'totalShots', 'shotsOnTarget', 'keyPasses',
    'successfulDribbles', 'dribbles', 'dribblesSuccess',
    'tackles', 'tacklesWon', 'interceptions',
    'totalDuelsWon', 'duelsWon', 'aerialDuelsWon',
    'accuratePasses', 'accuratePassesPercentage',
    'accurateCrosses', 'accurateCrossesPercentage',
    'accurateLongBalls', 'accurateLongBallsPercentage', 'totalLongBalls',
    'totalPasses', 'totalCross',
    'yellowCards', 'redCards',
    'minutesPlayed', 'appearances', 'rating',
    'expectedGoals', 'expectedAssists',
    'bigChancesCreated', 'bigChancesMissed',
    'clearances', 'blockedShots', 'ballRecovery', 'saves',
    'fouls', 'wasFouled', 'offsides', 'touches',
    'possessionLost', 'possessionWonAttThird', 'dispossessed',
    'penaltyGoals', 'penaltiesTaken',
    'goalConversionPercentage', 'penaltyConversion',
    'successfulDribblesPercentage', 'totalDuelsWonPercentage',
    'aerialDuelsWonPercentage', 'tacklesWonPercentage',
    'groundDuelsWon', 'groundDuelsWonPercentage',
    'errorLeadToGoal', 'errorLeadToShot',
    'headedGoals', 'freeKickGoal',
    'goalsFromInsideTheBox', 'goalsFromOutsideTheBox',
    'matchesStarted',
]

# Stats that are percentages / ratios / ratings — never compute p90
_P90_EXCLUDE = frozenset([
    'accuratePassesPercentage', 'accurateLongBallsPercentage',
    'accurateCrossesPercentage', 'successfulDribblesPercentage',
    'totalDuelsWonPercentage', 'aerialDuelsWonPercentage',
    'goalConversionPercentage', 'penaltyConversion', 'rating',
    'tacklesWonPercentage', 'groundDuelsWonPercentage',
])

_MIN_MINUTES = 450

# Map profile_category letters to the groups used in the UI
_CAT_TO_GROUP = {'G': 'GK', 'D': 'DEF', 'M': 'MID', 'F': 'FWD'}

@app.route('/player_scatter')
def player_scatter():
    competitions = get_all_competitions(OUTPUT_DIR)
    return render_template('player_scatter.html', competitions=competitions)

@app.route('/api/player_scatter_data')
def api_player_scatter_data():
    import glob as _glob
    import json as _json
    from data_loader import profile_category

    competition = request.args.get('competition', '').strip()
    country     = request.args.get('country',     '').strip()
    if not competition or not country:
        return jsonify({'error': 'competition and country are required'}), 400

    league_dir = os.path.join(OUTPUT_DIR, country, competition)
    if not os.path.isdir(league_dir):
        return jsonify({'error': 'competition directory not found'}), 404

    players   = []
    keys_seen = []

    for club_entry in sorted(os.scandir(league_dir), key=lambda e: e.name):
        if not club_entry.is_dir():
            continue
        players_dir = os.path.join(club_entry.path, 'Players')
        if not os.path.isdir(players_dir):
            continue
        for pf in sorted(os.scandir(players_dir), key=lambda e: e.name):
            if not pf.name.endswith('.json'):
                continue
            try:
                with open(pf.path, encoding='utf-8') as f:
                    raw = _json.load(f)
                profile  = raw.get('profile', {}).get('player', {})
                stat_raw = raw.get('statistics', {}).get('statistics', {})

                # Minimum-minutes filter
                minutes = stat_raw.get('minutesPlayed')
                if not isinstance(minutes, (int, float)) or minutes < _MIN_MINUTES:
                    continue

                pos_list = profile.get('positionsDetailed') or raw.get('positions', [])
                fallback = profile.get('position', pos_list[0] if pos_list else '')
                cat      = profile_category(pos_list, fallback)
                pos_grp  = _CAT_TO_GROUP.get(cat, 'MID')

                stats    = {}
                stats_p90 = {}
                for field in PLAYER_STAT_FIELDS:
                    val = stat_raw.get(field)
                    if val is not None and isinstance(val, (int, float)) and not isinstance(val, bool):
                        stats[field] = val
                        if field not in keys_seen:
                            keys_seen.append(field)
                        # Compute p90 for countable stats
                        if field not in _P90_EXCLUDE and minutes > 0:
                            stats_p90[field] = round((val / minutes) * 90, 2)

                player_name = raw.get('player_name', profile.get('name', ''))
                club_name   = raw.get('team', club_entry.name.replace('_', ' '))
                appearances = stat_raw.get('appearances')
                players.append({
                    'player':        player_name,
                    'player_slug':   pf.name,
                    'club':          club_name,
                    'position':      pos_grp,
                    'minutesPlayed': int(minutes),
                    'appearances':   appearances if isinstance(appearances, (int, float)) else 0,
                    'stats':         stats,
                    'stats_p90':     stats_p90,
                })
            except Exception:
                continue

    stat_keys = [k for k in PLAYER_STAT_FIELDS if k in keys_seen]
    return jsonify({'players': players, 'stat_keys': stat_keys})



# ── Match Stats ───────────────────────────────────────────────────────────────

import threading as _threading
import uuid as _uuid
from match_loader import (
    get_all_matches, get_match,
    get_player_match_history, compute_per90,
    MATCH_STATS_META, _DERIVED_STATS,
)

MATCH_OUTPUT_DIR = _resolve_dir('TACTICAL_MATCH_OUTPUT_DIR', '..', 'match_output')

# In-memory job registry for background scrape tasks
_scrape_jobs: dict = {}
_scrape_jobs_lock = _threading.Lock()


@app.route('/matches')
def matches():
    all_matches = get_all_matches(MATCH_OUTPUT_DIR)
    return render_template('matches.html', matches=all_matches)


@app.route('/match/<int:event_id>')
def match_detail(event_id: int):
    data = get_match(event_id, MATCH_OUTPUT_DIR)
    if data is None:
        return render_template('matches.html',
                               matches=get_all_matches(MATCH_OUTPUT_DIR),
                               error=f'Match {event_id} not found.'), 404
    return render_template('match_detail.html', match=data,
                           stat_meta=MATCH_STATS_META)


@app.route('/player_matches/<int:player_id>')
def player_matches(player_id: int):
    appearances = get_player_match_history(player_id, MATCH_OUTPUT_DIR)
    per90_data  = compute_per90(appearances)

    # Try to find the player's name + overall profile link
    player_name = ''
    profile_path = None
    if appearances:
        player_name = appearances[0]['player_info'].get('name', '')
    # Cross-lookup in output/ by player_id (scan JSON filenames for id fragment)
    for root, _dirs, files in os.walk(OUTPUT_DIR):
        for fname in files:
            if not fname.endswith('.json'):
                continue
            fpath = os.path.join(root, fname)
            try:
                import json as _json2
                with open(fpath, encoding='utf-8') as _f:
                    _d = _json2.load(_f)
                if str(_d.get('player_id', '')) == str(player_id):
                    rel = os.path.relpath(fpath, OUTPUT_DIR).replace('\\', '/')
                    parts = rel.split('/')
                    if len(parts) == 4:
                        profile_path = '/'.join(parts[:3]) + '/' + parts[3].replace('.json', '')
                    if not player_name:
                        player_name = _d.get('player_name', '')
                    break
            except Exception:
                pass
        if profile_path:
            break

    return render_template(
        'player_match_history.html',
        player_id=player_id,
        player_name=player_name,
        appearances=appearances,
        per90=per90_data,
        stat_meta=MATCH_STATS_META,
        derived_meta=_DERIVED_STATS,
        profile_path=profile_path,
    )


@app.route('/api/match_scrape', methods=['POST'])
def api_match_scrape():
    """Start a background bulk scrape. Returns {job_id}."""
    body   = request.get_json(silent=True) or {}
    inputs = body.get('inputs', [])
    if not inputs or not isinstance(inputs, list):
        return jsonify({'error': 'inputs must be a non-empty list'}), 400

    job_id = str(_uuid.uuid4())
    with _scrape_jobs_lock:
        _scrape_jobs[job_id] = {
            'status':    'pending',
            'logs':      [],
            'filepath':  None,
            'event_ids': [],
            'error':     None,
        }

    def _worker() -> None:
        try:
            import sys as _sys
            _coding = os.path.normpath(os.path.join(os.path.dirname(__file__), '..'))
            if _coding not in _sys.path:
                _sys.path.insert(0, _coding)
            from match_stats_scraper import run_bulk_scraper

            with _scrape_jobs_lock:
                _scrape_jobs[job_id]['status'] = 'running'
                _scrape_jobs[job_id]['logs'].append('Scraping started…')

            def _log(msg: str) -> None:
                with _scrape_jobs_lock:
                    _scrape_jobs[job_id]['logs'].append(msg)

            filepath, payload = run_bulk_scraper(inputs, log_fn=_log)

            event_ids = []
            if payload and 'matches' in payload:
                event_ids = [m.get('event_id') for m in payload['matches'] if m.get('event_id')]
            elif payload and 'event_id' in payload:
                event_ids = [payload['event_id']]

            with _scrape_jobs_lock:
                _scrape_jobs[job_id]['status']    = 'done'
                _scrape_jobs[job_id]['filepath']  = filepath
                _scrape_jobs[job_id]['event_ids'] = event_ids
                _scrape_jobs[job_id]['logs'].append(f'Done — saved to {os.path.basename(filepath or "")}')
        except Exception as exc:
            with _scrape_jobs_lock:
                _scrape_jobs[job_id]['status'] = 'error'
                _scrape_jobs[job_id]['error']  = str(exc)
                _scrape_jobs[job_id]['logs'].append(f'Error: {exc}')

    _threading.Thread(target=_worker, daemon=True).start()
    return jsonify({'job_id': job_id})


@app.route('/api/match_scrape_status/<job_id>')
def api_match_scrape_status(job_id: str):
    with _scrape_jobs_lock:
        job = _scrape_jobs.get(job_id)
    if job is None:
        return jsonify({'error': 'job not found'}), 404
    return jsonify({
        'status':    job['status'],
        'logs':      job['logs'],
        'filepath':  job['filepath'],
        'event_ids': job['event_ids'],
        'error':     job['error'],
    })


if __name__ == '__main__':
    app.run(debug=True, port=5000)
