"""
SofaScore Team Formation Scraper — core logic.

For a given team and league season, paginate through every match the team
played, fetch lineup data to extract the formation, and emit a JSON + Excel (.xlsx)
summary of formation usage.

Single usage (CLI)::
    python team_formation_scraper.py 2817 8 77559
    python team_formation_scraper.py "https://www.sofascore.com/football/team/barcelona/2817" "https://www.sofascore.com/football/tournament/spain/laliga/8#id:77559"

Output saved to Coding/match_output/
"""

import json
import os
import re
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

BASE_URL = "https://www.sofascore.com/api/v1"


# ── ID / URL parsers ──────────────────────────────────────────────────────────

def extract_team_id(text: str) -> int | None:
    """Parse a SofaScore team ID from a URL or plain integer string.

    Handles:
    - Plain integer: ``"2817"``
    - Team URL:      ``"https://www.sofascore.com/football/team/barcelona/2817"``

    Args:
        text: Raw user input.

    Returns:
        Parsed team ID, or ``None`` if not found.
    """
    text = text.strip()
    if text.isdigit():
        return int(text)
    # Last numeric segment in the URL path (before optional query/fragment)
    m = re.search(r"/team/[^/]+/(\d+)", text)
    if m:
        return int(m.group(1))
    return None


def extract_league_ids(text: str) -> tuple[int, int] | None:
    """Parse (unique_tournament_id, season_id) from a SofaScore league URL.

    Handles:
    - League URL: ``"https://…/tournament/spain/laliga/8#id:77559"``
    - Shorthand:  ``"8/77559"`` or ``"8 77559"``

    Args:
        text: Raw user input.

    Returns:
        ``(uniq_tid, season_id)`` tuple, or ``None`` if not parseable.
    """
    text = text.strip()

    # URL form: /tournament/.../8#id:77559  or  /tournament/.../8  + fragment
    m = re.search(r"/tournament/[^/]+/[^/]+/(\d+)(?:[^#]*#id:(\d+))?", text)
    if m:
        uniq_tid = int(m.group(1))
        season_id = int(m.group(2)) if m.group(2) else None
        if season_id:
            return uniq_tid, season_id
        # season missing from URL — can't proceed
        return None

    # Shorthand: "8/77559" or "8 77559"
    m = re.match(r"^(\d+)[/ ](\d+)$", text)
    if m:
        return int(m.group(1)), int(m.group(2))

    return None


# ── Selenium helpers ──────────────────────────────────────────────────────────

def _create_driver() -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()),
        options=options,
    )
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


def _fetch_json(driver: webdriver.Chrome, url: str, retries: int = 3) -> dict | None:
    for attempt in range(retries):
        try:
            driver.get(url)
            time.sleep(0.8)
            raw = driver.find_element("tag name", "pre").text
            return json.loads(raw)
        except Exception as exc:
            print(f"    Retry {attempt + 1} on {url} — {exc}")
            time.sleep(2 ** attempt)
    return None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _ts_to_iso(ts: int | None) -> str:
    if not ts:
        return ""
    return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d")


def _result_label(home_score: int | None, away_score: int | None, is_home: bool) -> str:
    if home_score is None or away_score is None:
        return "?"
    team_score = home_score if is_home else away_score
    opp_score  = away_score if is_home else home_score
    if team_score > opp_score:
        return "W"
    if team_score < opp_score:
        return "L"
    return "D"


# ── Core pipeline ─────────────────────────────────────────────────────────────

def fetch_team_season_events(
    driver: webdriver.Chrome,
    team_id: int,
    uniq_tid: int,
    season_id: int,
    log_fn: Callable[[str], None] = print,
) -> list[dict]:
    """Paginate through a league season and collect all events for *team_id*.

    Uses ``/unique-tournament/{uniq_tid}/season/{season_id}/events/last/{page}``
    which returns the most recent events first.  Pagination stops when
    ``hasNextPage`` is ``False``.

    Args:
        driver:     Active Selenium Chrome driver.
        team_id:    The SofaScore team ID to filter on.
        uniq_tid:   The unique tournament (league) ID.
        season_id:  The specific season ID.
        log_fn:     Callable for progress messages.

    Returns:
        List of raw event dicts from the API, filtered to only those where the
        target team appears as home or away team.
    """
    collected: list[dict] = []
    page = 0
    while True:
        url = (
            f"{BASE_URL}/unique-tournament/{uniq_tid}"
            f"/season/{season_id}/events/last/{page}"
        )
        data = _fetch_json(driver, url)
        if not data:
            log_fn(f"    [page {page}] Empty response — stopping.")
            break

        events = data.get("events", [])
        if not events:
            break

        for ev in events:
            home_id = ev.get("homeTeam", {}).get("id")
            away_id = ev.get("awayTeam", {}).get("id")
            if team_id in (home_id, away_id):
                collected.append(ev)

        log_fn(f"    Page {page}: {len(events)} league events, {len(collected)} for target team so far")

        if not data.get("hasNextPage", False):
            break

        page += 1
        time.sleep(0.5)

    return collected


def scrape_team_formations(
    team_id: int,
    uniq_tid: int,
    season_id: int,
    output_dir: str = "match_output",
    log_fn: Callable[[str], None] = print,
) -> tuple[str, str, dict]:
    """Full pipeline: collect events → fetch lineups → save JSON + Excel (.xlsx).

    Args:
        team_id:    SofaScore team ID.
        uniq_tid:   Unique tournament (league) ID.
        season_id:  Season ID.
        output_dir: Directory to write output files into.
        log_fn:     Callable for progress messages.

    Returns:
        ``(json_path, csv_path)`` — absolute paths to the saved files.

    Raises:
        RuntimeError: If no events are found or the team profile can't be fetched.
    """
    log_fn("=" * 55)
    log_fn("  SofaScore Team Formation Scraper")
    log_fn("=" * 55)

    driver = _create_driver()
    try:
        # ── 1. Fetch team name ─────────────────────────────────────────────
        log_fn(f"\n[1] Fetching team profile for ID {team_id}…")
        team_data = _fetch_json(driver, f"{BASE_URL}/team/{team_id}")
        team_name = (team_data or {}).get("team", {}).get("name", f"Team_{team_id}")
        log_fn(f"    Team: {team_name}")

        # ── 2. Collect all team events for this season ────────────────────
        log_fn(f"\n[2] Collecting league events (uniq_tid={uniq_tid}, season={season_id})…")
        raw_events = fetch_team_season_events(driver, team_id, uniq_tid, season_id, log_fn)
        if not raw_events:
            raise RuntimeError(
                f"No events found for team {team_id} in "
                f"tournament {uniq_tid} / season {season_id}."
            )

        # Collect team name and season name from first event
        first_ev = raw_events[0]
        season_name = first_ev.get("season", {}).get("name", str(season_id))

        log_fn(f"    Found {len(raw_events)} match(es) for {team_name} in {season_name}")

        # ── 3. Fetch lineups per event ─────────────────────────────────────
        log_fn(f"\n[3] Fetching lineups for {len(raw_events)} match(es)…")
        match_records: list[dict] = []
        formation_counter: Counter = Counter()

        for idx, ev in enumerate(raw_events, 1):
            event_id  = ev.get("id")
            home_id   = ev.get("homeTeam", {}).get("id")
            is_home   = (home_id == team_id)
            side      = "home" if is_home else "away"
            opp_side  = "away" if is_home else "home"

            home_team = ev.get("homeTeam", {}).get("name", "?")
            away_team = ev.get("awayTeam", {}).get("name", "?")
            opponent  = away_team if is_home else home_team

            home_score = ev.get("homeScore", {}).get("current")
            away_score = ev.get("awayScore", {}).get("current")
            date       = _ts_to_iso(ev.get("startTimestamp"))
            result     = _result_label(home_score, away_score, is_home)
            score_str  = f"{home_score}–{away_score}" if home_score is not None else "?–?"

            log_fn(f"    [{idx}/{len(raw_events)}] {date}  {home_team} vs {away_team}  (ID {event_id})")

            lineup_data = _fetch_json(driver, f"{BASE_URL}/event/{event_id}/lineups")
            team_formation = ""
            opp_formation  = ""
            if lineup_data:
                team_formation = lineup_data.get(side,     {}).get("formation", "")
                opp_formation  = lineup_data.get(opp_side, {}).get("formation", "")

            if team_formation:
                formation_counter[team_formation] += 1
            else:
                formation_counter["Unknown"] += 1

            match_records.append({
                "event_id":          event_id,
                "date":              date,
                "home_team":         home_team,
                "away_team":         away_team,
                "opponent":          opponent,
                "home_away":         "H" if is_home else "A",
                "team_formation":    team_formation or "Unknown",
                "opponent_formation": opp_formation or "Unknown",
                "home_score":        home_score,
                "away_score":        away_score,
                "score":             score_str,
                "result":            result,
            })

            time.sleep(0.8)  # gentle pacing

        # Sort chronologically (oldest first)
        match_records.sort(key=lambda r: r["date"])

    finally:
        driver.quit()

    # ── 4. Build output payload ────────────────────────────────────────────
    formation_summary = dict(formation_counter.most_common())

    payload = {
        "team_id":              team_id,
        "team_name":            team_name,
        "unique_tournament_id": uniq_tid,
        "season_id":            season_id,
        "season_name":          season_name,
        "scraped_at":           datetime.now(tz=timezone.utc).isoformat(),
        "total_matches":        len(match_records),
        "formation_summary":    formation_summary,
        "matches":              match_records,
    }

    # ── 5. Save files ──────────────────────────────────────────────────────
    os.makedirs(output_dir, exist_ok=True)

    safe_team = re.sub(r"[^\w\s-]", "", team_name).strip().replace(" ", "_")
    base_name = f"{safe_team}_formations_{uniq_tid}_{season_id}"

    json_path  = os.path.join(output_dir, base_name + ".json")
    xlsx_path  = os.path.join(output_dir, base_name + ".xlsx")

    with open(json_path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)

    # ── Excel workbook ──────────────────────────────────────────────────────
    wb = Workbook()

    # ── Sheet 1: Matches ────────────────────────────────────────────────
    ws_matches = wb.active
    ws_matches.title = "Matches"

    _xlsx_fields = [
        "event_id", "date", "home_team", "away_team",
        "home_away", "team_formation", "opponent_formation",
        "score", "result",
    ]
    _headers = [
        "Event ID", "Date", "Home Team", "Away Team",
        "H/A", "Formation", "Opp. Formation",
        "Score", "Result",
    ]

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center")
    win_fill     = PatternFill("solid", fgColor="C6EFCE")
    loss_fill    = PatternFill("solid", fgColor="FFC7CE")
    draw_fill    = PatternFill("solid", fgColor="FFEB9C")

    ws_matches.append(_headers)
    for cell in ws_matches[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    for row_idx, rec in enumerate(match_records, start=2):
        ws_matches.append([rec.get(f, "") for f in _xlsx_fields])
        result_val = rec.get("result", "")
        fill = win_fill if result_val == "W" else loss_fill if result_val == "L" else draw_fill if result_val == "D" else None
        if fill:
            for cell in ws_matches[row_idx]:
                cell.fill = fill

    # Auto-fit column widths
    _col_widths = [10, 12, 28, 28, 5, 12, 14, 8, 8]
    for col_idx, width in enumerate(_col_widths, start=1):
        ws_matches.column_dimensions[
            ws_matches.cell(row=1, column=col_idx).column_letter
        ].width = width

    # ── Sheet 2: Formation Summary ────────────────────────────────────
    ws_summary = wb.create_sheet("Formation Summary")
    ws_summary.append(["Formation", "Count"])
    for cell in ws_summary[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
    for formation, count in formation_summary.items():
        ws_summary.append([formation, count])
    ws_summary.column_dimensions["A"].width = 15
    ws_summary.column_dimensions["B"].width = 8

    wb.save(xlsx_path)

    # ── 6. Summary log ────────────────────────────────────────────────────
    log_fn(f"\n  Saved JSON  → {os.path.abspath(json_path)}")
    log_fn(f"  Saved Excel → {os.path.abspath(xlsx_path)}")
    log_fn(f"\n  Formation summary for {team_name}:")
    for formation, count in formation_summary.items():
        bar = "█" * count
        log_fn(f"    {formation:<12} {count:>3}×  {bar}")
    log_fn("\nDone!")

    return os.path.abspath(json_path), os.path.abspath(xlsx_path), payload


# ── CLI entry point ───────────────────────────────────────────────────────────

def main() -> None:
    """CLI runner.

    Usage::
        python team_formation_scraper.py <team_id_or_url> <uniq_tid/season_id_or_league_url>

    Examples::
        python team_formation_scraper.py 2817 8/77559
        python team_formation_scraper.py "https://www.sofascore.com/football/team/barcelona/2817" "https://www.sofascore.com/football/tournament/spain/laliga/8#id:77559"
    """
    args = sys.argv[1:]
    if len(args) < 2:
        print("Usage: python team_formation_scraper.py <team_id_or_url> <uniq_tid/season_id>")
        print("  e.g. python team_formation_scraper.py 2817 8/77559")
        sys.exit(1)

    team_id = extract_team_id(args[0])
    if team_id is None:
        print(f"ERROR: Cannot parse team ID from: {args[0]!r}")
        sys.exit(1)

    league = extract_league_ids(args[1])
    if league is None:
        print(f"ERROR: Cannot parse league/season IDs from: {args[1]!r}")
        sys.exit(1)

    uniq_tid, season_id = league

    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "match_output")
    scrape_team_formations(team_id, uniq_tid, season_id, output_dir)


if __name__ == "__main__":
    main()
